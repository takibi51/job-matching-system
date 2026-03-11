"""
求人マッチングシステム — Match
プロトタイプ(match-app)のUI/UXに準拠
+ Web求人自動取得機能を追加
"""

import streamlit as st
import pandas as pd
import io
import sys
import os
import html
import re
import time as _time
import urllib.parse
from datetime import datetime, timezone, timedelta

JST = timezone(timedelta(hours=9))


def _now_jst() -> datetime:
    """日本時間の現在時刻"""
    return datetime.now(JST)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from data_collector import (
    fetch_from_all_sources, parse_csv_upload, parse_text_input,
    generate_search_urls, SOURCE_NAMES, get_fetch_log, set_jooble_api_key, _log,
)
from scorer import rank_jobs, generate_search_queries, score_job
from candidate_loader import (
    load_all_candidates, load_candidate_upload, merge_candidate_uploads,
    extract_all_tags, SUPPORTED_EXTENSIONS, _detect_file_type,
)
from cache_manager import (
    save_jobs, search_jobs, get_all_jobs, get_stats, delete_old_jobs, clear_all,
    get_keywords, add_keyword, remove_keyword, get_enabled_keywords, toggle_keyword,
    update_keyword_status, add_collection_log, get_collection_logs,
    save_candidate, get_saved_candidates, delete_candidate,
    update_candidate, get_candidate_by_id, save_candidate_file, get_candidate_files,
    save_proposal, update_proposal_status, get_proposals, delete_proposal, PROPOSAL_STATUSES,
    save_interview_sheet, update_interview_sheet, get_interview_sheets, delete_interview_sheet,
    get_app_setting, set_app_setting,
    add_chat_message, get_chat_history, clear_chat_history,
    update_job_type, get_job_type_stats,
)
from ai_generator import (
    generate_scout_message, generate_concerns, generate_hireability,
    generate_proposal_resume, generate_interview_analysis,
    generate_progress_analysis, generate_job_improvements,
    evaluate_market_fit, MARKET_FIT_AXES,
    generate_chat_response, generate_candidate_profile,
    detect_chat_action,
)
from auth import check_password, check_session_timeout, render_logout_button, safe_url

# ============================================================
# Jooble APIキー読み込み（環境変数 → st.secrets → DB設定のフォールバック）
# ============================================================
_jooble_key = os.environ.get("JOOBLE_API_KEY", "")
if not _jooble_key:
    try:
        _jooble_key = st.secrets.get("api_keys", {}).get("jooble", "")
    except Exception:
        pass
if not _jooble_key:
    _jooble_key = get_app_setting("jooble_api_key", "")
if _jooble_key:
    set_jooble_api_key(_jooble_key)

# ============================================================
# 職域・職種プリセット（全タブ共通）
# ============================================================
_JOB_PRESETS = {
    "IT・エンジニア": [
        "フロントエンドエンジニア", "バックエンドエンジニア", "フルスタックエンジニア",
        "インフラエンジニア", "SRE", "DevOps", "社内SE", "情報システム",
        "データエンジニア", "データサイエンティスト", "AIエンジニア", "機械学習エンジニア",
        "iOSエンジニア", "Androidエンジニア", "モバイルエンジニア",
        "QAエンジニア", "テストエンジニア", "セキュリティエンジニア",
        "Webエンジニア", "サーバーサイドエンジニア", "組み込みエンジニア",
        "クラウドエンジニア", "ネットワークエンジニア", "DBA",
    ],
    "デザイン・クリエイティブ": [
        "Webデザイナー", "UIデザイナー", "UXデザイナー", "UI/UXデザイナー",
        "グラフィックデザイナー", "プロダクトデザイナー", "BXデザイナー",
        "アートディレクター", "クリエイティブディレクター",
        "動画クリエイター", "映像ディレクター", "イラストレーター",
        "DTPデザイナー", "エディトリアルデザイナー", "3Dデザイナー",
    ],
    "マーケティング・広報": [
        "Webマーケター", "デジタルマーケティング", "コンテンツマーケター",
        "SEOコンサルタント", "広告運用", "SNSマーケター", "CRM担当",
        "マーケティングマネージャー", "ブランドマネージャー",
        "広報・PR", "IR", "コピーライター", "コンテンツディレクター",
        "グロースハッカー", "プロダクトマーケティング",
    ],
    "営業・ビジネス": [
        "法人営業", "個人営業", "ルート営業", "新規開拓営業",
        "インサイドセールス", "フィールドセールス", "ソリューション営業",
        "カスタマーサクセス", "カスタマーサポート", "テクニカルサポート",
        "アカウントマネージャー", "セールスエンジニア",
        "事業開発", "アライアンス", "パートナーセールス",
    ],
    "企画・マネジメント": [
        "プロジェクトマネージャー", "プロダクトマネージャー",
        "Webディレクター", "プロデューサー",
        "事業企画", "経営企画", "商品企画", "サービス企画",
        "新規事業", "経営戦略",
    ],
    "管理・コーポレート": [
        "人事", "採用担当", "人事労務", "組織開発",
        "経理", "財務", "管理会計", "経営管理",
        "総務", "法務", "内部監査", "コンプライアンス",
        "秘書", "アシスタント", "オフィスマネージャー",
    ],
    "コンサルティング": [
        "ITコンサルタント", "戦略コンサルタント", "業務コンサルタント",
        "人事コンサルタント", "組織コンサルタント", "DXコンサルタント",
        "M&Aアドバイザー", "会計コンサルタント",
    ],
    "医療・ヘルスケア": [
        "看護師", "准看護師", "保健師", "助産師",
        "医師", "歯科医師", "薬剤師",
        "理学療法士", "作業療法士", "言語聴覚士",
        "臨床検査技師", "臨床工学技士", "放射線技師",
        "管理栄養士", "介護福祉士", "社会福祉士",
        "医療事務", "歯科衛生士", "歯科助手",
        "ケアマネージャー", "介護職", "看護助手",
        "柔道整復師", "鍼灸師",
    ],
    "教育・研修": [
        "講師", "教員", "塾講師", "研修講師",
        "キャリアアドバイザー", "キャリアコンサルタント",
        "教育企画", "研修企画",
    ],
    "制作・ライティング": [
        "Webライター", "編集者", "校正者", "コピーライター",
        "テクニカルライター", "シナリオライター",
        "カメラマン", "フォトグラファー", "翻訳者",
    ],
    "物流・製造": [
        "物流管理", "倉庫管理", "SCM",
        "生産管理", "品質管理", "品質保証",
        "製造オペレーター", "設備保全",
        "購買・調達", "バイヤー",
    ],
    "金融・不動産": [
        "ファイナンシャルプランナー", "証券アナリスト",
        "リスク管理", "融資審査", "資産運用",
        "不動産営業", "不動産管理", "用地仕入",
        "マンション管理", "プロパティマネジメント",
    ],
}

# 全職種リスト（フラット化）
_ALL_ROLES = []
for _roles in _JOB_PRESETS.values():
    _ALL_ROLES.extend(_roles)

# ============================================================
# 勤務地プルダウン選択肢
# ============================================================
_LOCATION_OPTIONS = [
    "全国", "リモート",
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県",
    "岐阜県", "静岡県", "愛知県", "三重県",
    "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
    "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県",
    "福岡県", "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
    "海外",
]

# ============================================================
# ページ設定
# ============================================================
st.set_page_config(
    page_title="Match - 人材マッチングシステム",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# 認証チェック（未認証時はログイン画面のみ表示）
# ============================================================
if not check_password():
    st.stop()
check_session_timeout()

# ============================================================
# 同期取得（Streamlit Cloud対応 — st.status()でリアルタイム進捗）
# ============================================================

def run_fetch_sync(kw_list, location, sources, status_container=None):
    """求人を同期的に取得（進捗をst.status()でリアルタイム表示）"""
    # Jooble APIキーを確実に再読み込み（環境変数 → st.secrets → DB のフォールバック）
    _key = os.environ.get("JOOBLE_API_KEY", "")
    if not _key:
        try:
            _key = st.secrets.get("api_keys", {}).get("jooble", "")
        except Exception:
            pass
    if not _key:
        _key = get_app_setting("jooble_api_key", "")
    if _key:
        set_jooble_api_key(_key)
        _log(f"🔑 APIキー読み込み済み: {_key[:8]}...")
    else:
        _log("⚠️ Jooble APIキーが見つかりません（環境変数 / st.secrets / DB すべて未設定）")

    start = _time.time()
    total_steps = len(kw_list) * len(sources)
    completed = 0
    all_jobs = []
    seen_urls = set()

    # キーワードのステータスを更新
    for kw in kw_list:
        try:
            update_keyword_status(kw, "fetching")
        except Exception:
            pass

    for si, source_name in enumerate(sources):
        for ki, kw in enumerate(kw_list):
            completed += 1
            pct = int((completed / max(total_steps, 1)) * 100)
            detail = f"📡 {source_name}: 「{kw}」({completed}/{total_steps}) — {len(all_jobs)}件取得済み"

            if status_container:
                status_container.update(label=f"求人取得中... {pct}%", state="running")
                status_container.write(detail)

            try:
                jobs = fetch_from_all_sources([kw], location, enabled_sources=[source_name])
                for job in jobs:
                    url = job.get("url", "")
                    if url and url not in seen_urls:
                        seen_urls.add(url)
                        all_jobs.append(job)
            except Exception as e:
                if status_container:
                    status_container.write(f"⚠️ {source_name} エラー: {e}")

    # DB保存
    elapsed = _time.time() - start
    if all_jobs:
        saved = save_jobs(all_jobs)
        add_collection_log(len(kw_list), len(all_jobs), saved, ",".join(sources), elapsed)
        result_msg = f"✅ {saved}件を新規保存（{elapsed:.1f}秒・{len(all_jobs)}件中）"
    else:
        add_collection_log(len(kw_list), 0, 0, ",".join(sources), elapsed)
        result_msg = "⚠️ 取得0件でした（データ管理タブの取得ログを確認してください）"

    # キーワードごとのDB件数を更新（実際の保存件数を渡す）
    _saved_count = saved if all_jobs else 0
    for kw in kw_list:
        try:
            # 複数キーワードの場合は均等割り（1キーワードなら全件数）
            _kw_count = _saved_count if len(kw_list) == 1 else 0
            update_keyword_status(kw, "done", jobs_found=_kw_count)
        except Exception:
            pass

    # フェッチログを表示（デバッグ情報）
    _fl = get_fetch_log()
    if status_container:
        if all_jobs:
            status_container.update(label=result_msg, state="complete")
        else:
            status_container.update(label=result_msg, state="error")
            # 0件の場合は取得ログを表示して原因を示す
            if _fl:
                status_container.write("--- 取得ログ（直近） ---")
                for _line in _fl[-20:]:
                    status_container.write(_line)

    return {"saved": saved if all_jobs else 0, "total": len(all_jobs), "elapsed": elapsed, "result": result_msg}


# ============================================================
# ユーティリティ
# ============================================================
def esc(text):
    return html.escape(str(text)) if text else ""


# 日本語サイトのドメイン（翻訳不要）
_JP_DOMAINS = [
    "mynavi.jp", "doda.jp", "en-japan.com", "type.jp", "rikunabi",
    "green-japan.com", "wantedly.com", "xn--pckua2a7gp15o89zb",
    "careerjet.jp", "hellowork", "en-gage.net", "indeed.co.jp",
    "jp.indeed.com", "r-agent.com", "job-medley.com", "bizreach.jp",
]


def _job_url(url: str) -> str:
    """求人URLを返す。英語サイトの場合はGoogle翻訳経由にする"""
    if not url:
        return ""
    url_lower = url.lower()
    # 日本語サイトはそのまま
    if any(d in url_lower for d in _JP_DOMAINS):
        return url
    # 日本語文字がURLに含まれていればそのまま
    if re.search(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]', url):
        return url
    # 英語サイト → Google翻訳経由
    return f"https://translate.google.com/translate?sl=en&tl=ja&u={urllib.parse.quote(url, safe='')}"


# 英語→日本語 タイトル変換テーブル（よくある求人タイトル用語）
_EN_JA_TITLE = [
    # 職種
    ("Senior", "シニア"), ("Junior", "ジュニア"), ("Lead", "リード"), ("Chief", "チーフ"),
    ("Principal", "プリンシパル"), ("Staff", "スタッフ"), ("Head of", "責任者 -"),
    ("Manager", "マネージャー"), ("Director", "ディレクター"), ("Vice President", "副社長"),
    ("Assistant", "アシスタント"), ("Associate", "アソシエイト"), ("Intern", "インターン"),
    # エンジニア系
    ("Software Engineer", "ソフトウェアエンジニア"), ("Frontend Engineer", "フロントエンドエンジニア"),
    ("Backend Engineer", "バックエンドエンジニア"), ("Full Stack Engineer", "フルスタックエンジニア"),
    ("Full-Stack Engineer", "フルスタックエンジニア"), ("Fullstack Engineer", "フルスタックエンジニア"),
    ("Product Engineer", "プロダクトエンジニア"), ("Data Engineer", "データエンジニア"),
    ("Machine Learning Engineer", "機械学習エンジニア"), ("ML Engineer", "MLエンジニア"),
    ("DevOps Engineer", "DevOpsエンジニア"), ("Site Reliability Engineer", "SRE"),
    ("QA Engineer", "QAエンジニア"), ("Security Engineer", "セキュリティエンジニア"),
    ("Platform Engineer", "プラットフォームエンジニア"), ("Infrastructure Engineer", "インフラエンジニア"),
    ("iOS Engineer", "iOSエンジニア"), ("Android Engineer", "Androidエンジニア"),
    ("Mobile Engineer", "モバイルエンジニア"), ("Embedded Engineer", "組み込みエンジニア"),
    ("Cloud Engineer", "クラウドエンジニア"),
    ("Engineer", "エンジニア"),
    # デザイン系
    ("Product Designer", "プロダクトデザイナー"), ("UX Designer", "UXデザイナー"),
    ("UI Designer", "UIデザイナー"), ("UX/UI Designer", "UX/UIデザイナー"),
    ("Web Designer", "Webデザイナー"), ("Graphic Designer", "グラフィックデザイナー"),
    ("Designer", "デザイナー"),
    # PM/PO系
    ("Product Manager", "プロダクトマネージャー"), ("Project Manager", "プロジェクトマネージャー"),
    ("Engineering Manager", "エンジニアリングマネージャー"),
    ("Product Owner", "プロダクトオーナー"), ("Program Manager", "プログラムマネージャー"),
    # ビジネス系
    ("Business Development", "事業開発"), ("Account Executive", "アカウントエグゼクティブ"),
    ("Account Manager", "アカウントマネージャー"), ("Sales Manager", "セールスマネージャー"),
    ("Sales Representative", "営業担当"), ("Sales Engineer", "セールスエンジニア"),
    ("Customer Success", "カスタマーサクセス"), ("Customer Support", "カスタマーサポート"),
    ("Marketing Manager", "マーケティングマネージャー"), ("Growth Manager", "グロースマネージャー"),
    ("Content Manager", "コンテンツマネージャー"), ("Community Manager", "コミュニティマネージャー"),
    ("Operations Manager", "オペレーションマネージャー"),
    # データ系
    ("Data Scientist", "データサイエンティスト"), ("Data Analyst", "データアナリスト"),
    ("Business Analyst", "ビジネスアナリスト"), ("Research Scientist", "リサーチサイエンティスト"),
    # HR・管理系
    ("Human Resources", "人事"), ("Recruiter", "リクルーター"), ("Talent Acquisition", "採用"),
    ("Office Manager", "オフィスマネージャー"), ("Executive Assistant", "エグゼクティブアシスタント"),
    # その他
    ("Consultant", "コンサルタント"), ("Specialist", "スペシャリスト"),
    ("Coordinator", "コーディネーター"), ("Administrator", "アドミニストレーター"),
    ("Architect", "アーキテクト"), ("Analyst", "アナリスト"), ("Developer", "デベロッパー"),
    ("Technician", "テクニシャン"), ("Supervisor", "スーパーバイザー"),
    ("Representative", "担当者"),
    # 補助語
    ("Remote", "リモート"), ("Part-time", "パートタイム"), ("Full-time", "フルタイム"),
    ("Contract", "契約"), ("Temporary", "派遣"),
]

_JP_CHAR_RE_APP = re.compile(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]')


def _translate_title(title: str) -> str:
    """英語タイトルを日本語に変換。既に日本語が含まれていればそのまま返す"""
    if not title:
        return title
    if _JP_CHAR_RE_APP.search(title):
        return title
    result = title
    for en, ja in _EN_JA_TITLE:
        result = re.sub(re.escape(en), ja, result, flags=re.IGNORECASE)
    return result


def _filter_jobs_by_category(jobs, selected_domain, selected_roles):
    """職域・職種でフィルタ。選択なしなら全件返す"""
    if not selected_domain or selected_domain == "すべて":
        return jobs
    roles = selected_roles if selected_roles else _JOB_PRESETS.get(selected_domain, [])
    if not roles:
        return jobs
    filtered = []
    for j in jobs:
        jtext = (j.get("title", "") + " " + j.get("description", "")).lower()
        if any(r.lower() in jtext for r in roles):
            filtered.append(j)
    return filtered


# エリアグループ（scorer.pyと同じロジック）
_LOCATION_AREA_MAP = {
    "北海道": ["北海道", "札幌"],
    "東京都": ["東京", "渋谷", "新宿", "港区", "千代田", "品川", "目黒", "中央区", "六本木", "丸の内", "大手町"],
    "神奈川県": ["神奈川", "横浜", "川崎"],
    "埼玉県": ["埼玉", "さいたま", "大宮"],
    "千葉県": ["千葉", "船橋", "幕張"],
    "愛知県": ["愛知", "名古屋"],
    "大阪府": ["大阪", "梅田", "難波", "心斎橋", "堺", "豊中", "吹田"],
    "京都府": ["京都"],
    "兵庫県": ["兵庫", "神戸", "三宮", "西宮", "尼崎"],
    "福岡県": ["福岡", "博多"],
    "広島県": ["広島"],
    "宮城県": ["宮城", "仙台"],
    "リモート": ["リモート", "在宅", "テレワーク", "フルリモート", "remote"],
}


def _filter_jobs_by_locations(jobs, selected_locations):
    """勤務地でフィルタ（OR条件）。'全国'のみ or 空ならフィルタしない"""
    if not selected_locations or selected_locations == ["全国"]:
        return jobs
    locs = [l for l in selected_locations if l != "全国"]
    if not locs:
        return jobs
    # 検索用キーワードリストを構築
    search_terms = []
    for loc in locs:
        search_terms.append(loc.replace("県", "").replace("府", "").replace("都", ""))
        for kw in _LOCATION_AREA_MAP.get(loc, []):
            search_terms.append(kw)
    search_terms = list(set(search_terms))

    filtered = []
    for j in jobs:
        jloc = j.get("location", "") + " " + j.get("title", "") + " " + j.get("description", "")
        if any(t in jloc for t in search_terms):
            filtered.append(j)
    return filtered


def _build_matching_excel(candidate, conditions, ranked_jobs):
    """営業用Excel（2シート）を生成して BytesIO を返す"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    buf = io.BytesIO()
    info = candidate.get("info", {}) if candidate else {}
    tags = candidate.get("tags", {}) if candidate else {}
    profile = generate_candidate_profile(candidate) if candidate else {}
    cand_name = candidate.get("name", "候補者") if candidate else "候補者"

    # --- Sheet1: 候補者プロフィール ---
    hard_skills = info.get("hard_skills", profile.get("hard_skills", []))
    soft_skills = info.get("soft_skills", profile.get("soft_skills", []))
    career_summary = info.get("career_summary", profile.get("career_summary", ""))
    market_score = info.get("market_score", profile.get("market_score", 0))
    market_reasons = info.get("market_reasons", profile.get("market_reasons", []))
    certifications = info.get("certifications", tags.get("certifications", []))
    industries = info.get("industries", tags.get("industries", []))
    languages = info.get("languages", tags.get("languages", []))
    experience_level = info.get("experience_level", tags.get("experience_level", ""))
    work_styles = info.get("work_styles", tags.get("work_styles", []))

    kws = conditions.get("keywords", []) if conditions else []
    sal_min = conditions.get("salary_min", "") if conditions else ""
    sal_max = conditions.get("salary_max", "") if conditions else ""
    loc = conditions.get("location", "") if conditions else ""

    profile_rows = [
        ("候補者名", cand_name),
        ("キャリア概要", career_summary),
        ("ハードスキル", "、".join(hard_skills) if isinstance(hard_skills, list) else str(hard_skills)),
        ("ソフトスキル", "、".join(soft_skills) if isinstance(soft_skills, list) else str(soft_skills)),
        ("保有資格", "、".join(certifications) if isinstance(certifications, list) else str(certifications)),
        ("業界経験", "、".join(industries) if isinstance(industries, list) else str(industries)),
        ("語学", "、".join(l.get("language", str(l)) if isinstance(l, dict) else str(l) for l in languages) if languages else ""),
        ("経験レベル", str(experience_level)),
        ("希望勤務形態", "、".join(work_styles) if isinstance(work_styles, list) else str(work_styles)),
        ("市場価値スコア", f"{market_score}点"),
        ("市場評価理由", "\n".join(f"・{r}" for r in market_reasons) if isinstance(market_reasons, list) else str(market_reasons)),
        ("", ""),
        ("＜検索条件＞", ""),
        ("キーワード", "、".join(kws)),
        ("希望年収", f"{sal_min}万〜{sal_max}万" if sal_min and sal_max else ""),
        ("希望勤務地", str(loc)),
    ]
    df_profile = pd.DataFrame(profile_rows, columns=["項目", "内容"])

    # --- Sheet2: マッチ求人一覧 ---
    job_rows = []
    for i, j in enumerate(ranked_jobs, 1):
        job_rows.append({
            "順位": i,
            "マッチ度": j.get("score", 0),
            "企業名": j.get("company", ""),
            "ポジション名": _translate_title(j.get("title", "")),
            "勤務地": j.get("location", ""),
            "年収": j.get("salary", ""),
            "ソース": j.get("source", ""),
            "マッチする理由": j.get("match_reasons", ""),
            "求人URL": _job_url(j.get("url", "")),
        })
    df_jobs = pd.DataFrame(job_rows) if job_rows else pd.DataFrame(
        columns=["順位", "マッチ度", "企業名", "ポジション名", "勤務地", "年収", "ソース", "マッチする理由", "求人URL"]
    )

    # --- Excel書き出し ---
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_profile.to_excel(writer, sheet_name="候補者プロフィール", index=False)
        df_jobs.to_excel(writer, sheet_name="マッチ求人一覧", index=False)

        # スタイリング
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for sheet_name in ["候補者プロフィール", "マッチ求人一覧"]:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 列幅調整
        ws1 = writer.sheets["候補者プロフィール"]
        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 80

        ws2 = writer.sheets["マッチ求人一覧"]
        col_widths = {"A": 6, "B": 8, "C": 25, "D": 35, "E": 15, "F": 15, "G": 12, "H": 50, "I": 50}
        for col, w in col_widths.items():
            ws2.column_dimensions[col].width = w

    buf.seek(0)
    return buf.getvalue()


def _get_label_thresholds():
    return get_app_setting("label_thresholds", {"recommended": 85, "good": 70})


def _score_badge(score):
    th = _get_label_thresholds()
    if score >= th.get("recommended", 85):
        cls, label = "score-high", "かなり相性が良い"
    elif score >= th.get("good", 70):
        cls, label = "score-high", "相性が良い"
    elif score >= 55:
        cls, label = "score-mid", "可能性あり"
    else:
        cls, label = "score-low", "要検討"
    return f'<span class="score-badge {cls}">{score}点 {label}</span>'


def _match_bar(score):
    cls = "match-fill-high" if score >= 60 else ("match-fill-mid" if score >= 30 else "match-fill-low")
    return f'<div class="match-bar"><div class="match-fill {cls}" style="width:{min(score,100)}%"></div></div>'


def _fit_tags(reasons_str):
    if not reasons_str:
        return ""
    tags = reasons_str.split(" / ")
    return " ".join(f'<span class="fit-tag">{esc(t)}</span>' for t in tags[:5])


def _cand_to_conditions(cand):
    c = cand.get("conditions", {})
    return {
        "keywords": c.get("keywords", []),
        "location": c.get("location", "大阪"),
        "salary_min": c.get("salary_min", 0),
        "salary_max": c.get("salary_max", 0),
        "age": c.get("age", 0),
        "prefer_kansai": c.get("prefer_kansai", True),
        "extra_keywords": c.get("extra_keywords", []),
    }


# ============================================================
# CSS（プロトタイプ準拠）
# ============================================================
st.markdown("""
<style>
    .stat-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 12px; padding: 1.2rem; color: white; text-align: center;
    }
    .stat-card-blue {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        border-radius: 12px; padding: 1.2rem; color: white; text-align: center;
    }
    .stat-card-green {
        background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
        border-radius: 12px; padding: 1.2rem; color: #1a3a2a; text-align: center;
    }
    .stat-num { font-size: 2rem; font-weight: 800; }
    .stat-label { font-size: 0.85rem; opacity: 0.9; }
    .job-card {
        border: 1px solid #e2e8f0; border-radius: 12px; padding: 1.2rem;
        margin-bottom: 0.8rem; background: white;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06); transition: all 0.2s;
    }
    .job-card:hover { border-color: #667eea; box-shadow: 0 4px 12px rgba(102,126,234,0.15); }
    .cand-card {
        border: 1px solid #e2e8f0; border-radius: 12px; padding: 1rem;
        margin-bottom: 0.6rem; background: white;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06); transition: all 0.2s;
    }
    .cand-card:hover { border-color: #43e97b; box-shadow: 0 4px 12px rgba(67,233,123,0.15); }
    .score-badge {
        display: inline-block; padding: 0.2rem 0.6rem; border-radius: 20px;
        font-weight: 700; font-size: 0.85rem;
    }
    .score-high { background: #d4edda; color: #155724; }
    .score-mid { background: #fff3cd; color: #856404; }
    .score-low { background: #f8d7da; color: #721c24; }
    .match-bar { background: #e9ecef; border-radius: 10px; height: 8px; overflow: hidden; margin: 0.3rem 0; }
    .match-fill { height: 100%; border-radius: 10px; transition: width 0.5s; }
    .match-fill-high { background: linear-gradient(90deg, #43e97b, #38f9d7); }
    .match-fill-mid { background: linear-gradient(90deg, #f9d423, #ff4e50); }
    .match-fill-low { background: linear-gradient(90deg, #ff4e50, #c62828); }
    .fit-tag {
        display: inline-block; padding: 0.15rem 0.5rem; margin: 0.1rem;
        border-radius: 12px; font-size: 0.75rem; background: #eef2ff; color: #3730a3;
    }
    .chat-msg {
        padding: 0.5rem 0.7rem; border-radius: 10px; margin: 0.2rem 0;
        font-size: 0.82rem; line-height: 1.4; word-break: break-word;
    }
    .chat-user { background: #eef2ff; text-align: right; }
    .chat-ai { background: #f0fdf4; border: 1px solid #d1fae5; }
    .ai-sidebar {
        position: sticky; top: 2rem;
        background: #fafbfc; border: 1px solid #e2e8f0; border-radius: 12px;
        padding: 0.8rem; max-height: 85vh; overflow-y: auto;
    }
    .ai-sidebar h4 { margin: 0 0 0.5rem; font-size: 0.95rem; color: #1F4E79; }
    .ai-quick-btn {
        display: inline-block; padding: 0.25rem 0.6rem; margin: 0.15rem;
        border-radius: 16px; font-size: 0.75rem; background: #eef2ff;
        color: #3730a3; border: 1px solid #c7d2fe; cursor: pointer;
        text-decoration: none;
    }
    .ai-quick-btn:hover { background: #c7d2fe; }
    .ai-chat-history {
        max-height: 40vh; overflow-y: auto; margin: 0.5rem 0;
        padding: 0.3rem; border-radius: 8px; background: white;
    }
    .empty-state {
        text-align: center; padding: 3rem; color: #a0aec0;
        border: 2px dashed #e2e8f0; border-radius: 16px; margin: 1rem 0;
    }
    .progress-bar-bg {
        background: #e9ecef; border-radius: 10px; height: 20px; overflow: hidden; margin: 0.5rem 0;
    }
    .progress-bar-fill {
        height: 100%; border-radius: 10px; background: linear-gradient(90deg, #667eea, #764ba2);
        transition: width 0.3s; text-align: center; color: white; font-size: 0.75rem; line-height: 20px;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# サイドバー
# ============================================================
st.sidebar.markdown("## 🎯 Match")
st.sidebar.caption("人材マッチングシステム")
render_logout_button()
st.sidebar.markdown("---")

tabs_def = {
    "candidate_search": "🔍 候補者検索",
    "job_search": "📋 求人検索",
    "interview": "📝 面談分析",
    "progress": "📊 提案管理",
    "data_import": "📦 データ取込",
}

if "current_page" not in st.session_state:
    st.session_state["current_page"] = "candidate_search"

for key, label in tabs_def.items():
    if st.sidebar.button(label, key=f"nav_{key}", use_container_width=True):
        st.session_state["current_page"] = key
        st.rerun()

page = st.session_state["current_page"]

# 統計
st.sidebar.markdown("---")
stats = get_stats()
saved_cands = get_saved_candidates()
jt_stats = get_job_type_stats()
contracted_count = jt_stats.get("contracted", 0)
web_count = jt_stats.get("web", 0)
st.sidebar.caption(f"👤 候補者: {len(saved_cands)}名")
st.sidebar.caption(f"📌 契約中求人: {contracted_count}件 / 🌐 Web掲載: {web_count}件")

# （同期取得のため、サイドバー進捗表示は不要）


# ============================================================
# 共通: AI右サイドパネル
# ============================================================
_TAB_QUICK_ACTIONS = {
    "candidateSearch": [
        ("🔍 求人を探して", "この候補者に合う求人を探してください"),
        ("📝 スカウト文", "スカウト文を提案してください"),
        ("⚠️ 懸念点", "懸念点を分析してください"),
        ("📊 決まりやすさ", "決まりやすさを教えてください"),
    ],
    "jobSearch": [
        ("👤 候補者を探して", "この求人に合う候補者を出してください"),
        ("💡 求人改善", "求人票を改善してください"),
    ],
    "interviewSheet": [
        ("📋 面談準備", "この候補者の面談準備をしてください"),
        ("❓ 質問案", "面談での深掘り質問を提案してください"),
    ],
    "proposals": [
        ("📊 進捗分析", "停滞している提案を分析してください"),
    ],
}


def render_ai_sidebar(tab_key, context=None):
    """右サイドパネルにAIアシスタントを表示"""
    st.markdown('<div class="ai-sidebar">', unsafe_allow_html=True)
    st.markdown("#### 🤖 AIアシスタント")

    # クイックアクションボタン
    quick_actions = _TAB_QUICK_ACTIONS.get(tab_key, [])
    for label, prompt in quick_actions:
        if st.button(label, key=f"ai_quick_{tab_key}_{label}", use_container_width=True):
            _handle_smart_chat(tab_key, prompt, context)

    st.markdown("---")

    # チャット履歴表示
    history = get_chat_history(tab_key, limit=15)
    if history:
        st.markdown('<div class="ai-chat-history">', unsafe_allow_html=True)
        for msg in history:
            role_cls = "chat-user" if msg["role"] == "user" else "chat-ai"
            avatar = "👤" if msg["role"] == "user" else "🤖"
            content_short = esc(msg["content"][:300]).replace("\n", "<br>")
            if len(msg["content"]) > 300:
                content_short += "..."
            st.markdown(f'<div class="chat-msg {role_cls}">{avatar} {content_short}</div>',
                        unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # チャット入力
    user_input = st.text_input(
        "メッセージ",
        placeholder="質問やリクエストを入力...",
        key=f"chat_input_{tab_key}",
        label_visibility="collapsed",
    )
    if st.button("送信", key=f"chat_send_{tab_key}", type="primary", use_container_width=True):
        if user_input and user_input.strip():
            _handle_smart_chat(tab_key, user_input.strip(), context)

    # クリア
    if history:
        if st.button("🗑️ 履歴クリア", key=f"chat_clear_{tab_key}", use_container_width=True):
            clear_chat_history(tab_key)
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)


def _handle_smart_chat(tab_key, message, context):
    """スマートチャット: インテント検出 → アクション実行 → 応答生成"""
    action = detect_chat_action(message, tab=tab_key, context=context)

    if action["action"] == "search_jobs":
        # 求人検索をトリガー
        kws = action["keywords"]
        if kws:
            search_query = " ".join(kws)
            results = search_jobs(search_query)
            if results:
                st.session_state[f"ai_search_results_{tab_key}"] = {
                    "type": "jobs",
                    "query": search_query,
                    "results": results,
                    "message": f"「{search_query}」で {len(results)}件 の求人が見つかりました。",
                }
                action["response"] = f"「{search_query}」で **{len(results)}件** の求人が見つかりました。メインエリアに表示しています。"
            else:
                action["response"] = f"「{search_query}」に一致する求人が見つかりませんでした。別のキーワードをお試しください。"

    elif action["action"] == "search_candidates":
        # 候補者検索をトリガー
        kws = action["keywords"]
        all_cands = get_saved_candidates()
        if kws and all_cands:
            matched = _filter_candidates_by_keywords(all_cands, kws)
            st.session_state[f"ai_search_results_{tab_key}"] = {
                "type": "candidates",
                "query": " ".join(kws),
                "results": matched,
                "message": f"「{', '.join(kws)}」に関連する候補者を {len(matched)}名 見つけました。",
            }
            action["response"] = f"「{', '.join(kws)}」に関連する候補者を **{len(matched)}名** 見つけました。メインエリアに表示しています。"
        elif all_cands:
            st.session_state[f"ai_search_results_{tab_key}"] = {
                "type": "candidates",
                "query": "",
                "results": all_cands,
                "message": f"登録済み候補者 {len(all_cands)}名 を表示しています。",
            }
            action["response"] = f"登録済みの候補者 **{len(all_cands)}名** をメインエリアに表示しています。"
        else:
            action["response"] = "候補者が登録されていません。「データ取込」から追加してください。"

    elif action["action"] == "sort_results":
        st.session_state[f"ai_sort_{tab_key}"] = action["sort"]

    # チャット履歴に保存
    add_chat_message(tab_key, "user", message)
    add_chat_message(tab_key, "ai", action["response"])
    st.rerun()


def _filter_candidates_by_keywords(candidates, keywords):
    """キーワードで候補者をフィルタリング"""
    if not keywords:
        return candidates
    matched = []
    for cand in candidates:
        text = " ".join([
            cand.get("name", ""),
            " ".join(str(v) for v in cand.get("info", {}).values()),
            " ".join(str(s) for s in cand.get("strengths", [])),
            " ".join(cand.get("conditions", {}).get("keywords", [])),
        ]).lower()
        score = sum(1 for kw in keywords if kw.lower() in text)
        if score > 0:
            matched.append((cand, score))
    matched.sort(key=lambda x: x[1], reverse=True)
    return [c for c, _ in matched]


def _render_ai_search_results(tab_key):
    """AIチャットからトリガーされた検索結果をメインカラムに表示"""
    key = f"ai_search_results_{tab_key}"
    if key not in st.session_state:
        return False

    data = st.session_state[key]
    st.markdown(f"### 🤖 AIアシスタントの検索結果")
    st.info(data["message"])

    if st.button("✕ 検索結果を閉じる", key=f"ai_close_{tab_key}"):
        del st.session_state[key]
        st.rerun()
        return False

    if data["type"] == "jobs":
        for i, job in enumerate(data["results"][:20]):
            jt_icon = "📌" if job.get("job_type") == "contracted" else "🌐"
            st.markdown(f"""
            <div class="job-card">
                <div style="font-size:1.05rem;font-weight:700;color:#1a202c;">
                    {jt_icon} {esc(_translate_title(job.get('title','')))}
                </div>
                🏢 {esc(job.get('company',''))} &nbsp;|&nbsp; 📍 {esc(job.get('location',''))}
                &nbsp;|&nbsp; 💰 {esc(job.get('salary',''))}
            </div>""", unsafe_allow_html=True)
            bc1, bc2 = st.columns(2)
            if bc1.button("📄 詳細", key=f"ai_job_{tab_key}_{i}"):
                show_job_popup(job, get_saved_candidates())
            url = _job_url(job.get("url", ""))
            if safe_url(url):
                bc2.markdown(f'<a href="{esc(url)}" target="_blank" style="display:inline-block;padding:0.4rem 1rem;border:1px solid #e2e8f0;border-radius:8px;text-decoration:none;color:#667eea;font-size:0.85rem;">🌐 求人ページ</a>', unsafe_allow_html=True)

    elif data["type"] == "candidates":
        for i, cand in enumerate(data["results"][:20]):
            mf = evaluate_market_fit(cand)
            star = "⭐ " if mf["has_star"] else ""
            kw_tags = " ".join(f'`{k}`' for k in cand.get("conditions", {}).get("keywords", [])[:5])
            st.markdown(f"""
            <div class="cand-card">
                {star}<strong>{esc(cand.get('name','候補者'))}</strong>
                <div style="margin-top:0.2rem;font-size:0.85rem;">{kw_tags}</div>
            </div>""", unsafe_allow_html=True)
            if st.button("👤 詳細", key=f"ai_cand_{tab_key}_{i}"):
                show_candidate_popup(cand)

    st.markdown("---")
    return True


# ============================================================
# 共通: ポップアップ
# ============================================================
@st.dialog("👤 候補者詳細", width="large")
def show_candidate_popup(cand):
    info = cand.get("info", {})
    conditions = cand.get("conditions", {})
    cid = cand.get("id", 0)

    # プロフィール自動生成（保存済みデータ優先）
    profile = generate_candidate_profile(cand)
    mf = evaluate_market_fit(cand)

    hard_skills = info.get("hard_skills", profile["hard_skills"])
    soft_skills = info.get("soft_skills", profile["soft_skills"])
    match_reasons = info.get("match_reasons", profile["match_reasons"])
    market_score = info.get("market_score", profile["market_score"])
    market_reasons = info.get("market_reasons", profile["market_reasons"])
    career_summary = info.get("career_summary", profile["career_summary"])
    personality_memo = info.get("personality_memo", profile["personality_memo"])
    negative_checks = info.get("negative_checks", profile["negative_checks"])

    # 拡張タグ情報
    tags = cand.get("tags", {})
    certifications = info.get("certifications", profile.get("certifications", tags.get("certifications", [])))
    industries = info.get("industries", profile.get("industries", tags.get("industries", [])))
    languages = info.get("languages", profile.get("languages", tags.get("languages", [])))
    management = info.get("management", profile.get("management", tags.get("management", {})))
    experience_level = info.get("experience_level", profile.get("experience_level", tags.get("experience_level", "")))
    achievements = info.get("achievements", profile.get("achievements", tags.get("achievements", [])))
    education = info.get("education", profile.get("education", tags.get("education", {})))
    work_styles = info.get("work_styles", profile.get("work_styles", tags.get("work_styles", [])))
    availability = info.get("availability", profile.get("availability", tags.get("availability", "")))
    career_change_reasons = info.get("career_change_reasons", profile.get("career_change_reasons", tags.get("career_change_reasons", [])))
    source_files = cand.get("source_files", [])

    # ===== ヘッダー =====
    st.markdown(f"### {cand.get('name', '候補者')}")
    st.caption(f"登録: {cand.get('created_at', '')[:10]}")

    # ===== メインコンテンツ + AIサイドバー =====
    main_col, ai_col = st.columns([3, 1])

    with ai_col:
        st.markdown("**🤖 AIパートナー**")
        if st.button("📝 スカウト文を提案", key=f"pop_scout_{cid}", use_container_width=True):
            st.session_state[f"pop_ai_{cid}"] = generate_scout_message(cand)
        if st.button("📋 企業向け提案文を生成", key=f"pop_resume_{cid}", use_container_width=True):
            st.session_state[f"pop_ai_{cid}"] = generate_proposal_resume(cand)
        if st.button("🎯 決定シナリオを考えて", key=f"pop_scenario_{cid}", use_container_width=True):
            st.session_state[f"pop_ai_{cid}"] = generate_hireability(cand)

        ai_input = st.text_input("自由に質問...", key=f"pop_ai_input_{cid}", label_visibility="collapsed",
                                 placeholder="自由に質問...")
        if st.button("▶", key=f"pop_ai_send_{cid}"):
            if ai_input and ai_input.strip():
                resp = generate_chat_response(ai_input, {"candidate": cand})
                st.session_state[f"pop_ai_{cid}"] = resp
        st.caption("↑ ボタンをクリックまたは質問を入力")

    with main_col:
        # マッチ理由
        st.markdown("#### 📋 マッチ理由")
        for r in match_reasons:
            st.markdown(f"• {r}")

        # スキル
        sk1, sk2 = st.columns(2)
        with sk1:
            st.markdown("#### 🔧 ハードスキル")
            hs_html = " ".join(f'<span class="fit-tag">{esc(s)}</span>' for s in hard_skills)
            st.markdown(hs_html or "情報なし", unsafe_allow_html=True)
        with sk2:
            st.markdown("#### 💡 ソフトスキル")
            ss_html = " ".join(
                f'<span class="fit-tag" style="background:#fef3c7;color:#92400e;">{esc(s)}</span>'
                for s in soft_skills
            )
            st.markdown(ss_html or "情報なし", unsafe_allow_html=True)

        # 資格・業界・語学・経験レベル（新規タグセクション）
        _has_extra_tags = any([certifications, industries, languages, experience_level, achievements])
        if _has_extra_tags:
            st.markdown("#### 📋 詳細プロフィール")
            _tag_cols = st.columns(3)
            with _tag_cols[0]:
                if certifications:
                    st.markdown("**📜 資格**")
                    for c in certifications[:6]:
                        cert_name = c if isinstance(c, str) else str(c)
                        st.markdown(f'<span class="fit-tag" style="background:#e0f2fe;color:#0369a1;">{esc(cert_name)}</span>', unsafe_allow_html=True)
                if experience_level:
                    st.markdown(f"**📊 経験レベル**: `{experience_level}`")
            with _tag_cols[1]:
                if industries:
                    st.markdown("**🏢 経験業界**")
                    for ind in industries[:5]:
                        st.markdown(f'<span class="fit-tag" style="background:#fce7f3;color:#9d174d;">{esc(ind)}</span>', unsafe_allow_html=True)
                if education and (education.get("level") or education.get("field")):
                    edu_parts = []
                    if education.get("level"):
                        edu_parts.append(education["level"])
                    if education.get("field"):
                        edu_parts.append(education["field"])
                    st.markdown(f"**🎓 学歴**: {' / '.join(edu_parts)}")
            with _tag_cols[2]:
                if languages:
                    st.markdown("**🌐 語学**")
                    for lang in languages[:4]:
                        lang_name = lang.get("language", "") if isinstance(lang, dict) else str(lang)
                        lang_level = lang.get("level", "") if isinstance(lang, dict) else ""
                        st.markdown(f'<span class="fit-tag" style="background:#ecfdf5;color:#065f46;">{esc(lang_name)} ({esc(lang_level)})</span>', unsafe_allow_html=True)
                if management and isinstance(management, dict) and management.get("has_experience"):
                    team_size = management.get("team_size", 0)
                    mgmt_label = f"{team_size}名規模" if team_size else "あり"
                    st.markdown(f"**👥 マネジメント**: `{mgmt_label}`")

        # 実績
        if achievements:
            st.markdown("#### 🏆 主な実績")
            for ach in achievements[:5]:
                ach_text = ach if isinstance(ach, str) else str(ach)
                st.markdown(f"• {ach_text}")

        # 市場決まりやすさ
        score_label = "注目" if market_score >= 75 else ("良好" if market_score >= 60 else "")
        badge_html = f' <span class="score-badge score-high">{score_label}</span>' if score_label else ""
        st.markdown(f"#### ⭐ 市場決まりやすさ: {market_score}%{badge_html}", unsafe_allow_html=True)
        for r in market_reasons:
            st.markdown(f"• {r}")

        # ネガティブチェック
        if negative_checks:
            st.markdown("#### ⚠️ ネガティブチェック")
            for nc in negative_checks:
                st.warning(nc)

        # 希望条件
        st.markdown("#### 🎯 希望条件")
        sal_min = conditions.get("salary_min", 0)
        sal_max = conditions.get("salary_max", 0)
        pref_parts = []
        if sal_min or sal_max:
            pref_parts.append(f"💰 {sal_min}万〜{sal_max}万円")
        if work_styles:
            for ws in work_styles:
                ws_icons = {"リモート": "🏠", "フレックス": "⏰", "時短": "🕐", "副業OK": "💼", "転勤なし": "📍"}
                pref_parts.append(f"{ws_icons.get(ws, '✓')} {ws}")
        elif conditions.get("remote") or any("リモート" in str(v) for v in conditions.values()):
            pref_parts.append("🏠 リモート希望")
        pref_parts.append(f"📍 {conditions.get('location', '未指定')}")
        if availability:
            pref_parts.append(f"📅 入社: {availability}")
        if career_change_reasons:
            pref_parts.append(f"💭 動機: {'・'.join(career_change_reasons[:2])}")
        st.markdown(" ・ ".join(pref_parts))

        # 職務要約
        st.markdown("#### 📝 職務要約")
        st.markdown(career_summary or "情報なし")

        # 人物タイプメモ（編集可能）
        st.markdown("#### 🧠 人物タイプメモ")
        edited_memo = st.text_area("人物タイプメモ", value=personality_memo or "", height=80,
                                   key=f"popup_memo_{cid}", label_visibility="collapsed")

        # ソースファイル情報
        if source_files:
            st.caption("📎 元ファイル: " + " / ".join(
                f"{sf.get('name', '')}" for sf in source_files if isinstance(sf, dict)
            ))

    # AI応答表示
    if st.session_state.get(f"pop_ai_{cid}"):
        with st.expander("🤖 AI応答", expanded=True):
            st.markdown(st.session_state[f"pop_ai_{cid}"])

    # ===== フッターボタン =====
    st.markdown("---")
    sheets = get_interview_sheets(cid) if cid else []
    proposals = get_proposals()
    cand_proposals = [p for p in proposals if p.get("candidate_id") == cid]

    f1, f2, f3, f4, f5 = st.columns(5)

    # 書類DL
    if f1.button("📄 書類DL", key=f"pop_dl_{cid}"):
        dl_text = f"# {cand.get('name', '候補者')}\n\n"
        dl_text += f"## 職務要約\n{career_summary}\n\n"
        dl_text += f"## ハードスキル\n{'、'.join(hard_skills)}\n\n"
        dl_text += f"## ソフトスキル\n{'、'.join(soft_skills)}\n\n"
        dl_text += f"## 希望条件\n{' / '.join(pref_parts)}\n\n"
        dl_text += f"## 人物タイプメモ\n{personality_memo}\n"
        st.session_state[f"pop_dl_data_{cid}"] = dl_text

    if st.session_state.get(f"pop_dl_data_{cid}"):
        st.download_button("⬇️ ダウンロード", st.session_state[f"pop_dl_data_{cid}"],
                           f"{cand.get('name', '候補者')}.md", "text/markdown",
                           key=f"pop_dl_btn_{cid}")

    # 面談シート確認
    if f2.button("📝 面談シート確認", key=f"pop_iv_{cid}"):
        st.session_state[f"pop_show_sheets_{cid}"] = not st.session_state.get(f"pop_show_sheets_{cid}", False)

    # 進捗変更
    with f3:
        if cand_proposals:
            current = cand_proposals[0].get("status", "提案済み")
            new_st = st.selectbox("進捗変更...", PROPOSAL_STATUSES,
                                  index=PROPOSAL_STATUSES.index(current) if current in PROPOSAL_STATUSES else 0,
                                  key=f"pop_status_{cid}")
            if new_st != current:
                if st.button("変更", key=f"pop_status_save_{cid}"):
                    update_proposal_status(cand_proposals[0]["id"], new_st)
                    st.success(f"「{new_st}」に変更")
                    st.rerun()
        else:
            st.caption("提案なし")

    # 提案用レジュメ生成
    if f4.button("📋 提案用レジュメ生成", key=f"pop_gen_resume_{cid}"):
        st.session_state[f"pop_ai_{cid}"] = generate_proposal_resume(cand)

    # 提案する
    if f5.button("📤 提案する", key=f"pop_propose_{cid}", type="primary"):
        st.session_state[f"pop_propose_mode_{cid}"] = not st.session_state.get(f"pop_propose_mode_{cid}", False)

    # メモ保存
    if edited_memo != (personality_memo or ""):
        if st.button("💾 メモを保存", key=f"pop_save_memo_{cid}", type="primary"):
            updated_info = {**info, "personality_memo": edited_memo}
            update_candidate(cid, info=updated_info)
            st.success("保存しました")
            st.rerun()

    # 面談シート表示
    if st.session_state.get(f"pop_show_sheets_{cid}"):
        st.markdown("---")
        if sheets:
            st.markdown("### 📝 面談シート")
            for sheet in sheets:
                created = sheet.get("created_at", "")[:16].replace("T", " ")
                tags = sheet.get("tags", [])
                tag_html = " ".join(f'<span class="fit-tag">#{esc(t)}</span>' for t in tags)
                st.markdown(f"**{created}** {tag_html}", unsafe_allow_html=True)
                edited_sheet = st.text_area("内容", value=sheet.get("sheet_content", ""),
                                            height=200, key=f"pop_sheet_{sheet['id']}")
                if edited_sheet != sheet.get("sheet_content", ""):
                    if st.button("シートを保存", key=f"pop_sheet_save_{sheet['id']}"):
                        update_interview_sheet(sheet["id"], edited_sheet, tags)
                        st.success("面談シートを更新しました")
                        st.rerun()
        else:
            st.info("面談シートがまだありません。「面談分析」タブで作成できます。")

    # 提案モード
    if st.session_state.get(f"pop_propose_mode_{cid}"):
        st.markdown("---")
        st.markdown("### 📤 提案先求人を選択")
        all_jobs_list = get_all_jobs(limit=50)
        if not all_jobs_list:
            st.info("求人がありません。「データ取込」タブで登録してください。")
        else:
            for ji, j in enumerate(all_jobs_list[:10]):
                jc1, jc2 = st.columns([4, 1])
                jt_icon = "📌" if j.get("job_type") == "contracted" else "🌐"
                jc1.markdown(f"{jt_icon} **{j.get('title', '')}** - {j.get('company', '')}")
                if jc2.button("提案", key=f"pop_prop_j_{cid}_{ji}"):
                    save_proposal(cid, j.get("url", ""), "提案済み", "")
                    st.success(f"「{j.get('title', '')}」への提案を登録しました")
                    st.session_state[f"pop_propose_mode_{cid}"] = False
                    st.rerun()


@st.dialog("📋 求人詳細", width="large")
def show_job_popup(job, candidates=None):
    st.markdown(f"### {_translate_title(job.get('title', '不明'))}")
    jt = job.get("job_type", "web")
    st.caption("📌 契約中" if jt == "contracted" else "🌐 Web掲載")

    j1, j2 = st.columns(2)
    with j1:
        st.markdown(f"🏢 **{job.get('company', '不明')}**")
        st.markdown(f"📍 {job.get('location', '不明')}")
        st.markdown(f"💰 {job.get('salary', '情報なし')}")
    with j2:
        url = _job_url(job.get("url", ""))
        if safe_url(url):
            st.markdown(f'<a href="{esc(url)}" target="_blank" style="display:inline-block;padding:0.4rem 1rem;border:1px solid #e2e8f0;border-radius:8px;text-decoration:none;color:#667eea;font-size:0.85rem;">🌐 求人ページを開く</a>', unsafe_allow_html=True)
        st.caption(f"ソース: {job.get('source', '')} / 更新: {job.get('updated_at', '')[:16]}")

    desc = job.get("description", "")
    if desc:
        with st.expander("説明文", expanded=True):
            st.markdown(desc[:600] + ("..." if len(desc) > 600 else ""))

    if candidates:
        st.markdown("---")
        st.markdown("**マッチする候補者:**")
        cand_scores = []
        for c in candidates:
            cond = _cand_to_conditions(c)
            sc, reasons = score_job(job, cond)
            if sc >= 10:
                cand_scores.append((c, sc, reasons))
        cand_scores.sort(key=lambda x: x[1], reverse=True)
        for c, sc, reasons in cand_scores[:5]:
            mf = evaluate_market_fit(c)
            star = "⭐ " if mf["has_star"] else ""
            st.markdown(f"{_score_badge(sc)} **{star}{esc(c.get('name', ''))}**", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("**🤖 AIアシスタント**")
    jurl = job.get("url", "none")
    if st.button("💡 求人改善提案", key=f"pop_jimp_{jurl[:30]}"):
        st.session_state["pop_job_ai"] = generate_job_improvements(job)
    if st.session_state.get("pop_job_ai"):
        st.markdown(st.session_state["pop_job_ai"])


# ############################################################
# タブ1: 候補者検索（プロトタイプ準拠）
# ############################################################
if page == "candidate_search":
    st.markdown("## 🔍 候補者検索")
    st.caption("候補者の条件にマッチする求人を検索します")

    col_sel, col_upload = st.columns([2, 1])
    with col_sel:
        if saved_cands:
            cand_options = ["-- 選択してください --"] + [
                f"{c['name']}（{c.get('created_at','')[:10]}）" for c in saved_cands
            ]
            sel = st.selectbox("保存済み候補者から選択", cand_options, key="cs_select")
            sel_idx = cand_options.index(sel) - 1 if sel != cand_options[0] else -1
        else:
            st.info("候補者が登録されていません。「📦 データ取込」から追加してください。")
            sel_idx = -1
    with col_upload:
        st.markdown("**クイックアップロード**（複数可）")
        ext_list = list(SUPPORTED_EXTENSIONS.keys())
        quick_uploads = st.file_uploader(
            "候補者ファイル", type=[e.lstrip(".") for e in ext_list],
            key="cs_quick_upload", label_visibility="collapsed",
            accept_multiple_files=True,
        )
        if quick_uploads:
            parsed_files = []
            for uf in quick_uploads:
                file_bytes = uf.read()
                cand_data = load_candidate_upload(file_bytes, uf.name)
                if cand_data:
                    doc_type = _detect_file_type(uf.name, " ".join(str(v) for v in cand_data.get("info", {}).values()))
                    cand_data["_file_type"] = doc_type
                    parsed_files.append(cand_data)
            if parsed_files:
                merged = merge_candidate_uploads(parsed_files) if len(parsed_files) > 1 else parsed_files[0]
                tags = extract_all_tags(
                    " ".join(str(v) for v in merged.get("info", {}).values()),
                    merged.get("info", {})
                )
                merged["tags"] = tags
                st.session_state["quick_cand"] = merged
                st.session_state["quick_cand_files"] = [uf.name for uf in quick_uploads]
                st.success(f"{len(parsed_files)}件読み取り完了")

    active_cand = None
    conditions = None

    if quick_uploads and "quick_cand" in st.session_state:
        active_cand = st.session_state["quick_cand"]
        conditions = active_cand.get("conditions", {})
        _default_name = quick_uploads[0].name.rsplit(".", 1)[0] if quick_uploads else "候補者"
        save_name = st.text_input("候補者名", value=_default_name, key="cs_save_name")
        if st.button("この候補者を保存", key="cs_save"):
            save_candidate(save_name, active_cand.get("info", {}),
                           active_cand.get("strengths", []), conditions,
                           tags=active_cand.get("tags", {}))
            st.success(f"「{save_name}」を保存しました")
            st.rerun()
    elif sel_idx >= 0:
        active_cand = saved_cands[sel_idx]
        conditions = _cand_to_conditions(active_cand)

    if active_cand and conditions:
        mf_active = evaluate_market_fit(active_cand)
        star_txt = "⭐ " if mf_active["has_star"] else ""
        kw_tags = " ".join(f'`{k}`' for k in conditions.get("keywords", []))
        st.markdown(f"**{star_txt}{active_cand.get('name', '候補者')}** — {kw_tags}")
        if st.button("👤 候補者詳細を開く", key="cs_cand_popup"):
            show_candidate_popup(active_cand)

        with st.expander("⚙️ 検索条件を調整", expanded=False):
            ac1, ac2, ac3 = st.columns(3)
            salary_min = ac1.number_input("最低年収(万)", value=conditions.get("salary_min", 300), step=10, key="cs_smin")
            salary_max = ac2.number_input("最高年収(万)", value=conditions.get("salary_max", 600), step=10, key="cs_smax")
            age_val = ac3.number_input("年齢", value=max(conditions.get("age", 30), 18), min_value=18, max_value=70, key="cs_age")
            _loc_default = conditions.get("location", "全国")
            _loc_defaults = [_loc_default] if _loc_default and _loc_default != "全国" else []
            loc_vals = st.multiselect("勤務地（複数選択可・OR条件）", _LOCATION_OPTIONS, default=_loc_defaults, key="cs_loc")
            kw_str = st.text_area("キーワード（改行区切り）",
                                  value="\n".join(conditions.get("keywords", [])), height=60, key="cs_kw")
            kws = [k.strip() for k in kw_str.split("\n") if k.strip()]
            loc_val = loc_vals[0] if len(loc_vals) == 1 else ("全国" if not loc_vals else loc_vals[0])
            conditions = {
                "keywords": kws, "location": loc_val,
                "salary_min": salary_min, "salary_max": salary_max,
                "age": age_val, "prefer_kansai": True,
                "extra_keywords": conditions.get("extra_keywords", []),
                "_locations": loc_vals,  # 複数勤務地（内部用）
            }

        st.markdown("---")

        # 2カラムレイアウト: メイン結果 + AIサイドパネル
        _cs_main, _cs_ai = st.columns([3, 1])

        with _cs_main:
            # AIトリガーの検索結果
            _render_ai_search_results("candidateSearch")

            # フィルタ行: 求人種別 + 職域 + 職種
            _cf1, _cf2 = st.columns(2)
            jt_filter = _cf1.radio("求人種別", ["すべて", "📌 契約中", "🌐 Web掲載"], horizontal=True, key="cs_jt")
            jt_val = None
            if "契約中" in jt_filter:
                jt_val = "contracted"
            elif "Web" in jt_filter:
                jt_val = "web"

            _cf3, _cf4 = st.columns(2)
            cs_domain = _cf3.selectbox("職域で絞り込み", ["すべて"] + list(_JOB_PRESETS.keys()), key="cs_domain")
            cs_roles = []
            if cs_domain != "すべて":
                cs_roles = _cf4.multiselect("職種で絞り込み", _JOB_PRESETS.get(cs_domain, []), key="cs_roles")

            if stats["total_jobs"] == 0:
                st.markdown('<div class="empty-state"><h3>求人データがまだありません</h3>'
                            '<p>「📦 データ取込」からデータを追加してください</p></div>',
                            unsafe_allow_html=True)
            else:
                search_query = " ".join(conditions.get("keywords", []) + conditions.get("extra_keywords", [])[:3])
                matched_jobs = search_jobs(search_query) if search_query.strip() else get_all_jobs(limit=300, job_type=jt_val)
                if jt_val and matched_jobs:
                    matched_jobs = [j for j in matched_jobs if j.get("job_type", "web") == jt_val]
                # 職域・職種フィルタ
                matched_jobs = _filter_jobs_by_category(matched_jobs, cs_domain, cs_roles)
                # 勤務地フィルタ（複数選択OR条件）
                matched_jobs = _filter_jobs_by_locations(matched_jobs, conditions.get("_locations", []))

                if matched_jobs:
                    ranked = rank_jobs(matched_jobs, conditions)
                    fc1, fc2, fc3 = st.columns(3)
                    min_score = fc1.slider("最低スコア", 0, 100, 0, key="cs_fscore")
                    sort_opt = fc2.selectbox("並べ替え", ["マッチ度順", "年収順（高→低）", "新着順"], key="cs_sort")
                    require_salary = fc3.checkbox("年収あり", value=False, key="cs_fsal")

                    filtered = [j for j in ranked if j.get("score", 0) >= min_score
                                and (not require_salary or j.get("salary", "").strip())]
                    if sort_opt == "年収順（高→低）":
                        def _sk(j):
                            nums = re.findall(r'(\d+)', j.get("salary", "").replace(",", ""))
                            return max([int(n) for n in nums]) if nums else 0
                        filtered.sort(key=_sk, reverse=True)
                    elif sort_opt == "新着順":
                        filtered.sort(key=lambda x: x.get("updated_at", ""), reverse=True)

                    st.markdown(f"**{len(filtered)}件**マッチ（全{stats['total_jobs']:,}件中）")

                    for i, job in enumerate(filtered[:50]):
                        score = job.get("score", 0)
                        jt_icon = "📌" if job.get("job_type") == "contracted" else "🌐"
                        st.markdown(f"""
                        <div class="job-card">
                            {_score_badge(score)} <span style="color:#a0aec0;font-size:0.8rem;">{jt_icon} #{i+1}</span>
                            {_match_bar(score)}
                            <div style="font-size:1.05rem;font-weight:700;color:#1a202c;margin:0.3rem 0;">
                                {esc(_translate_title(job.get('title','不明')))}
                            </div>
                            🏢 <strong>{esc(job.get('company',''))}</strong>
                            &nbsp;|&nbsp; 📍 {esc(job.get('location',''))}
                            &nbsp;|&nbsp; 💰 {esc(job.get('salary',''))}
                            <div style="margin-top:0.3rem;">{_fit_tags(job.get('match_reasons',''))}</div>
                        </div>""", unsafe_allow_html=True)

                        qa1, qa2, qa3 = st.columns([1, 1, 1])
                        if qa1.button("📄 求人詳細", key=f"cs_jpop_{i}"):
                            show_job_popup(job, saved_cands)
                        if active_cand and active_cand.get("id"):
                            if qa2.button("📊 提案登録", key=f"cs_prop_{i}"):
                                save_proposal(active_cand["id"], job.get("url", ""), "提案済み", "")
                                st.success("提案を登録しました")
                                st.rerun()
                        url = _job_url(job.get("url", ""))
                        if safe_url(url):
                            qa3.markdown(f'<a href="{esc(url)}" target="_blank" style="display:inline-block;padding:0.4rem 1rem;border:1px solid #e2e8f0;border-radius:8px;text-decoration:none;color:#667eea;font-size:0.85rem;">🌐 求人ページ</a>', unsafe_allow_html=True)

                    with st.expander(f"📊 テーブル表示（{len(filtered)}件）"):
                        df = pd.DataFrame([{
                            "順位": i, "スコア": j.get("score", 0),
                            "種別": "契約中" if j.get("job_type") == "contracted" else "Web",
                            "求人タイトル": _translate_title(j.get("title", "")), "企業名": j.get("company", ""),
                            "勤務地": j.get("location", ""), "年収": j.get("salary", ""),
                            "ソース": j.get("source", ""), "URL": _job_url(j.get("url", "")),
                        } for i, j in enumerate(filtered, 1)])
                        st.dataframe(df, hide_index=True, use_container_width=True)

                        # --- Excel出力（2シート: 候補者情報 + 求人一覧） ---
                        _dl1, _dl2 = st.columns(2)
                        with _dl1:
                            csv_buf = io.StringIO()
                            df.to_csv(csv_buf, index=False, encoding="utf-8-sig")
                            st.download_button("📄 CSV DL", csv_buf.getvalue(), "マッチング結果.csv", "text/csv", key="cs_csv_dl")
                        with _dl2:
                            xlsx_buf = _build_matching_excel(active_cand, conditions, filtered)
                            cand_name = active_cand.get("name", "候補者") if active_cand else "候補者"
                            st.download_button(
                                "📊 Excel DL（営業用）", xlsx_buf,
                                f"マッチング_{cand_name}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="cs_xlsx_dl",
                            )
                else:
                    st.info("条件に一致する求人がありません。")

        with _cs_ai:
            _cs_chat_ctx = {"candidate": active_cand if active_cand else None, "tab": "candidateSearch"}
            render_ai_sidebar("candidateSearch", _cs_chat_ctx)

    else:
        _cs_main2, _cs_ai2 = st.columns([3, 1])
        with _cs_main2:
            st.markdown("#### 手動検索")
            manual_kw = st.text_input("検索キーワード", placeholder="例: Webデザイナー 大阪", key="cs_manual_kw")
            if manual_kw.strip() and stats["total_jobs"] > 0:
                results = search_jobs(manual_kw)
                if results:
                    st.markdown(f"**{len(results)}件** ヒット")
                    for i, job in enumerate(results[:20]):
                        jt_icon = "📌" if job.get("job_type") == "contracted" else "🌐"
                        st.markdown(f"""
                        <div class="job-card">
                            {jt_icon} <strong>{esc(_translate_title(job.get('title','')))}</strong> - {esc(job.get('company',''))}
                            &nbsp;|&nbsp; 📍 {esc(job.get('location',''))} &nbsp;|&nbsp; 💰 {esc(job.get('salary',''))}
                        </div>""", unsafe_allow_html=True)
            # AIトリガーの検索結果
            _render_ai_search_results("candidateSearch")
        with _cs_ai2:
            render_ai_sidebar("candidateSearch", {"tab": "candidateSearch"})


# ############################################################
# タブ2: 求人検索
# ############################################################
elif page == "job_search":
    st.markdown("## 📋 求人検索")
    st.caption("求人にマッチする候補者を探します")

    # フィルタ行
    _jf1, _jf2 = st.columns(2)
    js_jt = _jf1.radio("求人種別", ["すべて", "📌 契約中", "🌐 Web掲載"], horizontal=True, key="js_jt_filter")
    js_jt_val = None
    if "契約中" in js_jt:
        js_jt_val = "contracted"
    elif "Web" in js_jt:
        js_jt_val = "web"

    _jf3, _jf4 = st.columns(2)
    js_domain = _jf3.selectbox("職域で絞り込み", ["すべて"] + list(_JOB_PRESETS.keys()), key="js_domain")
    js_roles = []
    if js_domain != "すべて":
        js_roles = _jf4.multiselect("職種で絞り込み", _JOB_PRESETS.get(js_domain, []), key="js_roles")

    js_locs = st.multiselect("勤務地で絞り込み（複数選択可・OR条件）", _LOCATION_OPTIONS, default=[], key="js_locs")

    job_search_kw = st.text_input("求人を検索", placeholder="職種・企業名・キーワード", key="js_kw")
    if job_search_kw.strip():
        job_results = search_jobs(job_search_kw)
    else:
        job_results = get_all_jobs(limit=300, job_type=js_jt_val)
    if js_jt_val and job_results:
        job_results = [j for j in job_results if j.get("job_type", "web") == js_jt_val]
    # 職域・職種フィルタ
    job_results = _filter_jobs_by_category(job_results, js_domain, js_roles)
    # 勤務地フィルタ（複数選択OR条件）
    job_results = _filter_jobs_by_locations(job_results, js_locs)

    selected_job_for_chat = None

    # 2カラムレイアウト: メイン結果 + AIサイドパネル
    _js_main, _js_ai = st.columns([3, 1])

    with _js_main:
        # AIトリガーの検索結果
        _render_ai_search_results("jobSearch")

        if not job_results:
            st.markdown('<div class="empty-state"><h3>条件に一致する求人がありません</h3></div>', unsafe_allow_html=True)
        elif not saved_cands:
            st.markdown('<div class="empty-state"><h3>候補者が登録されていません</h3></div>', unsafe_allow_html=True)
        else:
            st.markdown(f"**{len(job_results)}件**の求人")
            for i, job in enumerate(job_results[:30]):
                jt_icon = "📌" if job.get("job_type") == "contracted" else "🌐"
                st.markdown(f"""
                <div class="job-card">
                    <div style="font-size:1.05rem;font-weight:700;color:#1a202c;">
                        {jt_icon} {esc(_translate_title(job.get('title','')))}
                    </div>
                    🏢 {esc(job.get('company',''))} &nbsp;|&nbsp; 📍 {esc(job.get('location',''))}
                    &nbsp;|&nbsp; 💰 {esc(job.get('salary',''))}
                </div>""", unsafe_allow_html=True)

                jqa1, jqa2 = st.columns(2)
                if jqa1.button("📄 詳細 & マッチ候補者", key=f"js_jpop_{i}"):
                    show_job_popup(job, saved_cands)
                url = _job_url(job.get("url", ""))
                if safe_url(url):
                    jqa2.markdown(f'<a href="{esc(url)}" target="_blank" style="display:inline-block;padding:0.4rem 1rem;border:1px solid #e2e8f0;border-radius:8px;text-decoration:none;color:#667eea;font-size:0.85rem;">🌐 求人ページ</a>', unsafe_allow_html=True)

                cand_scores = []
                for cand in saved_cands:
                    cond = _cand_to_conditions(cand)
                    sc, reasons = score_job(job, cond)
                    if sc >= 10:
                        cand_scores.append((cand, sc, reasons))
                cand_scores.sort(key=lambda x: x[1], reverse=True)
                if cand_scores:
                    top3 = cand_scores[:3]
                    tcols = st.columns(len(top3))
                    for ti, (cand, sc, reasons) in enumerate(top3):
                        with tcols[ti]:
                            mf_c = evaluate_market_fit(cand)
                            star = "⭐" if mf_c["has_star"] else ""
                            st.markdown(f"{_score_badge(sc)} {star} **{esc(cand.get('name',''))}**",
                                        unsafe_allow_html=True)
                            if st.button("👤 詳細", key=f"js_cpop_{i}_{cand.get('id',ti)}"):
                                show_candidate_popup(cand)
                if i == 0:
                    selected_job_for_chat = job

    with _js_ai:
        _js_chat_ctx = {"job": selected_job_for_chat if job_results else None, "tab": "jobSearch"}
        render_ai_sidebar("jobSearch", _js_chat_ctx)


# ############################################################
# タブ3: 面談分析
# ############################################################
elif page == "interview":
    st.markdown("## 📝 面談分析")
    st.caption("面談内容をAIが構造化し、候補者特性を抽出します")

    _iv_main, _iv_ai = st.columns([3, 1])

    # 右サイドパネル: AIアシスタント
    with _iv_ai:
        render_ai_sidebar("interviewSheet", {"tab": "interviewSheet"})

    with _iv_main:
        iv_tabs = st.tabs(["✏️ 新規作成", "📋 シート一覧"])

    with iv_tabs[0]:
        if not saved_cands:
            st.markdown('<div class="empty-state"><h3>候補者を先に登録してください</h3></div>',
                        unsafe_allow_html=True)
        else:
            iv_cand_options = [f"{c['name']}（ID:{c['id']}）" for c in saved_cands]
            iv_sel = st.selectbox("対象候補者", iv_cand_options, key="iv_cand_sel")
            iv_sel_idx = iv_cand_options.index(iv_sel) if iv_sel else 0
            iv_cand = saved_cands[iv_sel_idx]

            if st.button("👤 候補者詳細", key="iv_cand_popup"):
                show_candidate_popup(iv_cand)

            st.markdown("---")
            raw_input = st.text_area(
                "面談内容を入力（自由記述）", height=250,
                placeholder="面談で聞き取った内容をそのまま入力してください。",
                key="iv_raw_input",
            )
            tag_input = st.text_input("タグ（カンマ区切り）", placeholder="例: 医療事務, 即日可", key="iv_tags")

            if st.button("面談シートを生成", type="primary", key="iv_generate"):
                if not raw_input.strip():
                    st.error("面談内容を入力してください")
                else:
                    tags = [t.strip() for t in tag_input.split(",") if t.strip()] if tag_input else []
                    auto_patterns = {
                        "即日可": ["即日", "すぐ"], "大阪希望": ["大阪"], "東京希望": ["東京"],
                        "リモート希望": ["リモート", "在宅"], "管理職志向": ["管理職", "マネジメント"],
                        "医療系": ["医療", "クリニック", "病院"], "IT系": ["エンジニア", "IT"],
                    }
                    for tag_name, keywords in auto_patterns.items():
                        if tag_name not in tags and any(kw in raw_input for kw in keywords):
                            tags.append(tag_name)

                    lines = raw_input.strip().split("\n")
                    sections = {"職歴・経験": [], "転職理由・動機": [], "希望条件": [],
                                "スキル・資格": [], "人物像・印象": [], "その他": []}
                    section_kw = {
                        "職歴・経験": ["現職", "前職", "経験", "勤務", "職歴"],
                        "転職理由・動機": ["転職理由", "動機", "退職"],
                        "希望条件": ["希望", "年収", "勤務地", "条件"],
                        "スキル・資格": ["スキル", "資格", "免許"],
                        "人物像・印象": ["人柄", "印象", "性格"],
                    }
                    cur_sec = "その他"
                    for line in lines:
                        lc = line.strip()
                        if not lc:
                            continue
                        for sec, kws in section_kw.items():
                            if any(kw in lc for kw in kws):
                                cur_sec = sec
                                break
                        sections[cur_sec].append(lc)

                    parts = [f"# 面談シート: {iv_cand.get('name', '候補者')}",
                             f"作成日: {_now_jst().strftime('%Y-%m-%d %H:%M')}", ""]
                    for sec_name, sec_lines in sections.items():
                        if sec_lines:
                            parts.append(f"## {sec_name}")
                            for sl in sec_lines:
                                parts.append(f"- {sl.lstrip('-').lstrip('・').strip()}")
                            parts.append("")
                    if tags:
                        parts.append("## タグ")
                        parts.append(", ".join(f"#{t}" for t in tags))
                    sheet_content = "\n".join(parts)

                    ai_result = generate_interview_analysis(raw_input, iv_cand)
                    for at in ai_result.get("tags", []):
                        tc = at.lstrip("#")
                        if tc not in tags:
                            tags.append(tc)
                    sheet_content += "\n\n" + ai_result.get("report", "")

                    sheet_id = save_interview_sheet(iv_cand["id"], raw_input, sheet_content, tags)

                    # 候補者プロフィールを面談内容で自動充実化
                    enriched = generate_candidate_profile(iv_cand, interview_text=raw_input)
                    updated_info = dict(iv_cand.get("info", {}))
                    for key in ["hard_skills", "soft_skills", "match_reasons", "market_score",
                                "market_reasons", "career_summary", "personality_memo", "negative_checks"]:
                        if enriched.get(key) and not updated_info.get(key):
                            updated_info[key] = enriched[key]
                    update_candidate(iv_cand["id"], info=updated_info)

                    st.success(f"面談シートを保存し、候補者プロフィールを自動充実化しました（ID: {sheet_id}）")
                    st.markdown("---")
                    st.markdown("### 生成されたシート")
                    st.markdown(sheet_content)
                    if tags:
                        tag_html = " ".join(f'<span class="fit-tag">#{esc(t)}</span>' for t in tags)
                        st.markdown(tag_html, unsafe_allow_html=True)

    with iv_tabs[1]:
        all_sheets = get_interview_sheets()
        if not all_sheets:
            st.markdown('<div class="empty-state"><h3>面談シートがまだありません</h3></div>',
                        unsafe_allow_html=True)
        else:
            cand_map = {c["id"]: c["name"] for c in saved_cands}
            if saved_cands:
                iv_filter = st.selectbox("候補者でフィルタ", ["全員"] + [c["name"] for c in saved_cands], key="iv_filter")
            else:
                iv_filter = "全員"

            for sheet in all_sheets:
                cand_name = cand_map.get(sheet["candidate_id"], f"#{sheet['candidate_id']}")
                if iv_filter != "全員" and cand_name != iv_filter:
                    continue
                tags = sheet.get("tags", [])
                tag_html = " ".join(f'<span class="fit-tag">#{esc(t)}</span>' for t in tags)
                created = sheet.get("created_at", "")[:16].replace("T", " ")
                st.markdown(f"""
                <div class="cand-card">
                    <strong>{esc(cand_name)}</strong>
                    <span style="color:#a0aec0;margin-left:1rem;font-size:0.8rem;">{created}</span>
                    <div style="margin-top:0.3rem;">{tag_html}</div>
                </div>""", unsafe_allow_html=True)
                with st.expander(f"📋 詳細 - {cand_name}"):
                    edited_content = st.text_area("内容", value=sheet.get("sheet_content", ""),
                                                  height=250, key=f"iv_edit_{sheet['id']}",
                                                  label_visibility="collapsed")
                    iv_bc1, iv_bc2 = st.columns(2)
                    if edited_content != sheet.get("sheet_content", ""):
                        if iv_bc1.button("💾 保存", key=f"iv_save_{sheet['id']}"):
                            update_interview_sheet(sheet["id"], edited_content, tags)
                            st.success("更新しました")
                            st.rerun()
                    if iv_bc2.button("🗑️ 削除", key=f"iv_del_{sheet['id']}"):
                        delete_interview_sheet(sheet["id"])
                        st.rerun()



# ############################################################
# タブ4: 提案管理
# ############################################################
elif page == "progress":
    st.markdown("## 📊 提案管理")
    st.caption("候補者×求人の提案状況を管理します")

    _pr_main, _pr_ai = st.columns([3, 1])
    with _pr_ai:
        render_ai_sidebar("proposals", {"tab": "proposals"})

    with _pr_main:
        proposals = get_proposals()
        cand_map = {c["id"]: c for c in saved_cands}

        status_counts = {s: 0 for s in PROPOSAL_STATUSES}
        for p in proposals:
            s = p.get("status", "提案済み")
            if s in status_counts:
                status_counts[s] += 1

        if proposals:
            colors = ["#667eea", "#4facfe", "#43e97b", "#38f9d7", "#f9d423", "#ff6b6b", "#ee5a24", "#6c5ce7"]
            pipeline_html = '<div style="display:flex;gap:4px;margin-bottom:1rem;">'
            for i, (status, count) in enumerate(status_counts.items()):
                color = colors[i % len(colors)]
                pipeline_html += f'<div style="flex:1;text-align:center;padding:0.5rem;border-radius:8px;background:{color}20;border:2px solid {color};"><div style="font-size:1.5rem;font-weight:800;color:{color};">{count}</div><div style="font-size:0.7rem;color:#4a5568;">{esc(status)}</div></div>'
            pipeline_html += '</div>'
            st.markdown(pipeline_html, unsafe_allow_html=True)

        with st.expander("➕ 新しい提案を登録"):
            if saved_cands and stats["total_jobs"] > 0:
                pr_cand_opts = [f"{c['name']}（ID:{c['id']}）" for c in saved_cands]
                pr_cand_sel = st.selectbox("候補者", pr_cand_opts, key="pr_cand")
                pr_cand_idx = pr_cand_opts.index(pr_cand_sel)
                pr_cand_id = saved_cands[pr_cand_idx]["id"]
                pr_job_kw = st.text_input("求人を検索", key="pr_job_kw")
                pr_jobs = search_jobs(pr_job_kw) if pr_job_kw.strip() else get_all_jobs(limit=50)
                if pr_jobs:
                    pr_job_opts = [f"{j.get('title','不明')} - {j.get('company','')}" for j in pr_jobs[:30]]
                    pr_job_sel = st.selectbox("求人", pr_job_opts, key="pr_job")
                    pr_job = pr_jobs[pr_job_opts.index(pr_job_sel)]
                    pr_memo = st.text_input("メモ", key="pr_memo")
                    if st.button("提案を登録", type="primary", key="pr_save"):
                        save_proposal(pr_cand_id, pr_job.get("url", ""), "提案済み", pr_memo)
                        st.success("登録しました")
                        st.rerun()
            else:
                st.info("候補者と求人の両方が必要です")

        st.markdown("---")

        if not proposals:
            st.markdown('<div class="empty-state"><h3>提案がまだありません</h3></div>', unsafe_allow_html=True)
        else:
            pf1, pf2 = st.columns(2)
            pr_filter_status = pf1.multiselect("ステータス", PROPOSAL_STATUSES, default=PROPOSAL_STATUSES, key="pr_fstatus")
            if saved_cands:
                pr_filter_cand = pf2.selectbox("候補者", ["全員"] + [c["name"] for c in saved_cands], key="pr_fcand")
            else:
                pr_filter_cand = "全員"

            filtered_proposals = []
            for p in proposals:
                if p.get("status", "") not in pr_filter_status:
                    continue
                cn = cand_map.get(p["candidate_id"], {}).get("name", f"#{p['candidate_id']}")
                if pr_filter_cand != "全員" and cn != pr_filter_cand:
                    continue
                filtered_proposals.append(p)

            st.markdown(f"**{len(filtered_proposals)}件**")

            for p in filtered_proposals:
                cand_info = cand_map.get(p["candidate_id"], {})
                cand_name = cand_info.get("name", f"#{p['candidate_id']}")
                status = p.get("status", "提案済み")
                updated = p.get("updated_at", "")[:16].replace("T", " ")

                from cache_manager import _get_conn
                job_info = None
                job_url = p.get("job_url", "")
                if job_url:
                    conn = _get_conn()
                    row = conn.execute("SELECT * FROM jobs WHERE url = ?", (job_url,)).fetchone()
                    if row:
                        job_info = dict(row)
                job_title = job_info.get("title", "不明") if job_info else "不明"

                s_colors = {"提案済み": "#667eea", "カジュアル面談": "#4facfe", "一次面接": "#43e97b",
                            "二次面接": "#38f9d7", "三次面接": "#f9d423", "内定": "#ff6b6b",
                            "内定承諾": "#ee5a24", "決定": "#6c5ce7"}
                s_color = s_colors.get(status, "#667eea")

                st.markdown(f"""
                <div class="job-card">
                    <span style="display:inline-block;padding:0.2rem 0.7rem;border-radius:12px;background:{s_color}20;color:{s_color};font-weight:700;font-size:0.85rem;">{esc(status)}</span>
                    <span style="color:#a0aec0;font-size:0.8rem;margin-left:0.5rem;">{updated}</span>
                    <div style="margin-top:0.4rem;">
                        <strong>👤 {esc(cand_name)}</strong> → <strong>📋 {esc(job_title)}</strong>
                    </div>
                    {f'<div style="color:#718096;font-size:0.85rem;">📝 {esc(p.get("memo",""))}</div>' if p.get("memo") else ''}
                </div>""", unsafe_allow_html=True)

                with st.expander(f"⚙️ 操作 - {cand_name}"):
                    pu1, pu2 = st.columns(2)
                    current_idx = PROPOSAL_STATUSES.index(status) if status in PROPOSAL_STATUSES else 0
                    new_status = pu1.selectbox("ステータス", PROPOSAL_STATUSES, index=current_idx, key=f"pr_ns_{p['id']}")
                    new_memo = pu2.text_input("メモ", value=p.get("memo", ""), key=f"pr_nm_{p['id']}")
                    new_next = st.text_input("次のアクション", value=p.get("next_action", ""), key=f"pr_na_{p['id']}")

                    uc1, uc2, uc3 = st.columns(3)
                    if uc1.button("更新", type="primary", key=f"pr_upd_{p['id']}"):
                        update_proposal_status(p["id"], new_status, new_memo, new_next)
                        st.success("更新しました")
                        st.rerun()
                    if uc2.button("🗑️ 削除", key=f"pr_del_{p['id']}"):
                        delete_proposal(p["id"])
                        st.rerun()
                    if uc3.button("🤖 進捗分析", key=f"pr_ai_{p['id']}"):
                        analysis = generate_progress_analysis({**p, "job_title": job_title}, cand_info or None)
                        st.session_state[f"pr_analysis_{p['id']}"] = analysis
                    if st.session_state.get(f"pr_analysis_{p['id']}"):
                        st.markdown(st.session_state[f"pr_analysis_{p['id']}"])



# ############################################################
# タブ5: データ取込（新規追加機能）
# ############################################################
elif page == "data_import":
    st.markdown("## 📦 データ取込")
    st.caption("求人データの取り込み・候補者の登録を行います")

    dc1, dc2, dc3 = st.columns(3)
    dc1.markdown(f'<div class="stat-card"><div class="stat-num">{contracted_count}</div>'
                 f'<div class="stat-label">📌 契約中求人</div></div>', unsafe_allow_html=True)
    dc2.markdown(f'<div class="stat-card-blue"><div class="stat-num">{web_count}</div>'
                 f'<div class="stat-label">🌐 Web掲載求人</div></div>', unsafe_allow_html=True)
    dc3.markdown(f'<div class="stat-card-green"><div class="stat-num">{len(saved_cands)}</div>'
                 f'<div class="stat-label">👤 候補者数</div></div>', unsafe_allow_html=True)

    st.markdown("")
    dm_tabs = st.tabs(["🌐 Web求人取得", "📌 契約中求人登録", "👤 候補者登録", "🗄️ データ管理"])

    # --- Web求人取得 ---
    with dm_tabs[0]:
        st.markdown("### Web求人の自動取得")
        st.caption("求人サイトからキーワード検索で自動取得します。")

        # --- fetchingのまま放置されたキーワードをリセット ---
        _stale_kws = [k for k in get_keywords() if k.get("fetch_status") == "fetching"]
        for _sk in _stale_kws:
            try:
                update_keyword_status(_sk["keyword"], "pending")
            except Exception:
                pass

        # --- 半日自動更新チェック ---
        _AUTO_REFRESH_HOURS = 12
        _last_auto = get_app_setting("last_auto_fetch_at")
        _now_ts = _now_jst()
        if _last_auto:
            try:
                _last_dt = datetime.fromisoformat(_last_auto)
                _hours_since = (_now_ts - _last_dt).total_seconds() / 3600
            except (ValueError, TypeError):
                _hours_since = _AUTO_REFRESH_HOURS + 1
        else:
            _hours_since = _AUTO_REFRESH_HOURS + 1

        _auto_kws = get_enabled_keywords()
        _auto_refresh_needed = _hours_since >= _AUTO_REFRESH_HOURS and _auto_kws

        # --- API設定 ---
        with st.expander("🔑 API設定（求人データの取得に必要）", expanded=not _jooble_key):
            if _jooble_key:
                st.success("Jooble APIキー: 設定済み")
            else:
                st.warning("⚠️ Jooble APIキーが未設定です。クラウド環境では求人データを取得できません。")
                st.markdown("""
**設定手順（無料・5分で完了）:**
1. [Jooble API登録ページ](https://jooble.org/api/about) にアクセス
2. メールアドレスを入力してAPIキーを取得
3. 下のフォームにキーを貼り付けて保存
""")
            with st.form("api_key_form"):
                _new_jooble_key = st.text_input(
                    "Jooble APIキー",
                    value=_jooble_key or "",
                    type="password",
                    placeholder="APIキーを入力",
                )
                if st.form_submit_button("保存"):
                    if _new_jooble_key.strip():
                        set_app_setting("jooble_api_key", _new_jooble_key.strip())
                        set_jooble_api_key(_new_jooble_key.strip())
                        st.success("APIキーを保存しました。")
                        st.rerun()
            # 接続テスト
            if _jooble_key:
                if st.button("🧪 API接続テスト", key="dm_api_test"):
                    with st.status("Jooble API接続テスト中...", expanded=True) as _tc:
                        import requests as _req
                        _test_ok = False
                        for _ep in [f"https://jooble.org/api/{_jooble_key}"]:
                            try:
                                _tc.write(f"テスト: {_ep[:45]}...")
                                _test_resp = _req.post(
                                    _ep,
                                    json={"keywords": "エンジニア", "location": "Japan", "page": 1},
                                    headers={"Content-Type": "application/json"},
                                    timeout=20,
                                )
                                _tc.write(f"  ステータス: {_test_resp.status_code} / {len(_test_resp.text)}文字")
                                if _test_resp.status_code == 200:
                                    _test_data = _test_resp.json()
                                    _tc.write(f"  totalCount: {_test_data.get('totalCount', '?')}")
                                    _tc.write(f"  jobs: {len(_test_data.get('jobs', []))}件")
                                    if _test_data.get("jobs"):
                                        _tc.write(f"  例: {_test_data['jobs'][0].get('title', '?')}")
                                        _test_ok = True
                                        break
                                else:
                                    _tc.write(f"  レスポンス: {_test_resp.text[:300]}")
                            except Exception as _e:
                                _tc.write(f"  エラー: {_e}")
                        if _test_ok:
                            _tc.update(label="API接続成功 — 求人データ取得可能", state="complete")
                        else:
                            _tc.update(label="API接続失敗 — 求人データが取得できません", state="error")

        st.markdown("**取得ソース:**")
        enabled_sources = []
        src_cols = st.columns(len(SOURCE_NAMES))
        for i, name in enumerate(SOURCE_NAMES):
            _src_default = name in SOURCE_NAMES[:1]  # Joobleをデフォルト有効
            if src_cols[i].checkbox(name, value=True, key=f"dm_src_{name}"):
                enabled_sources.append(name)

        registered_kws = get_enabled_keywords()
        if registered_kws:
            st.markdown("**有効キーワード:** " + ", ".join([f"「{kw['keyword']}」" for kw in registered_kws[:10]]))

        # 職域×職種のプリセット辞書（グローバル定義を使用）

        with st.expander("🔑 キーワード管理", expanded=True):
            kw_tab1, kw_tab2 = st.tabs(["📝 手動追加", "👤 候補者から追加"])

            with kw_tab1:
                # フリーワード追加
                with st.form("dm_add_kw"):
                    kc1, kc2 = st.columns([3, 1])
                    new_kw = kc1.text_input("フリーワード", placeholder="例: Webデザイナー")
                    new_kw_loc = kc2.selectbox("勤務地", _LOCATION_OPTIONS, index=0, key="dm_kw_loc")
                    _auto_start = st.checkbox("登録と同時に求人取得を開始する", value=True, key="dm_auto_fetch_free")
                    if st.form_submit_button("追加"):
                        if new_kw.strip():
                            _loc_val = "" if new_kw_loc == "全国" else new_kw_loc
                            if add_keyword(new_kw.strip(), _loc_val):
                                st.success(f"「{new_kw}」を追加")
                                if _auto_start:
                                    with st.status("求人取得中...", expanded=True) as _sc:
                                        run_fetch_sync([new_kw.strip()], _loc_val, list(SOURCE_NAMES), status_container=_sc)
                                    set_app_setting("last_auto_fetch_at", _now_jst().isoformat())
                                st.rerun()

                st.markdown("---")
                st.markdown("**職域から選択して追加**")
                _preset_domain = st.selectbox(
                    "職域を選択", ["-- 選択してください --"] + list(_JOB_PRESETS.keys()),
                    key="dm_preset_domain"
                )
                if _preset_domain != "-- 選択してください --":
                    _preset_roles = _JOB_PRESETS[_preset_domain]
                    _existing_kws_preset = {kw["keyword"] for kw in get_keywords()}
                    _available_roles = [r for r in _preset_roles if r not in _existing_kws_preset]

                    if _available_roles:
                        _sel_all = st.checkbox("すべて選択", key=f"dm_preset_all_{_preset_domain}")
                        _selected_roles = st.multiselect(
                            f"{_preset_domain} の職種",
                            _available_roles,
                            default=_available_roles if _sel_all else [],
                            key=f"dm_preset_roles_{_preset_domain}",
                        )
                        if _selected_roles:
                            st.caption(f"**追加予定（{len(_selected_roles)}件）:** " +
                                       ", ".join(f"「{r}」" for r in _selected_roles))
                            _auto_start_preset = st.checkbox("登録と同時に求人取得を開始する", value=True, key="dm_auto_fetch_preset")
                            if st.button("選択した職種をキーワード登録", type="primary", key="dm_preset_add"):
                                _added = 0
                                _added_kws = []
                                for r in _selected_roles:
                                    if add_keyword(r, ""):
                                        _added += 1
                                        _added_kws.append(r)
                                st.success(f"{_added}件のキーワードを追加しました")
                                if _auto_start_preset and _added_kws:
                                    with st.status("求人取得中...", expanded=True) as _sc:
                                        run_fetch_sync(_added_kws, "", list(SOURCE_NAMES), status_container=_sc)
                                    set_app_setting("last_auto_fetch_at", _now_jst().isoformat())
                                st.rerun()
                    else:
                        st.info(f"{_preset_domain} の職種はすべて登録済みです")

            with kw_tab2:
                st.caption("登録候補者のスキル・職種からキーワードを自動提案します")
                _cands_for_kw = get_saved_candidates()
                if _cands_for_kw:
                    _existing_kws = {kw["keyword"] for kw in get_keywords()}

                    # 候補者選択（プルダウン）
                    _cand_options = ["全ての候補者"] + [c["name"] for c in _cands_for_kw]
                    _sel_cand = st.selectbox("候補者を選択", _cand_options, key="dm_sug_cand_sel")

                    if _sel_cand == "全ての候補者":
                        _target_cands = _cands_for_kw
                    else:
                        _target_cands = [c for c in _cands_for_kw if c["name"] == _sel_cand]

                    # カテゴリ別にキーワード候補を抽出
                    _sug_categories = {
                        "💼 職種・職域": {},
                        "🔧 スキル": {},
                        "📜 資格": {},
                        "🏢 業界": {},
                    }
                    for c in _target_cands:
                        tags = c.get("tags", {})
                        info = c.get("info", {})
                        # スキル（候補者ローダーの全キーワードを取得）
                        for s in tags.get("skills", []):
                            if s not in _existing_kws:
                                _sug_categories["🔧 スキル"].setdefault(s, []).append(c["name"])
                        # 職種
                        for s in tags.get("job_domains", []):
                            if s not in _existing_kws:
                                _sug_categories["💼 職種・職域"].setdefault(s, []).append(c["name"])
                        # 現在・希望職種
                        for key in ["desired_position", "current_position", "希望職種", "職種",
                                     "役職", "ポジション"]:
                            val = info.get(key, "")
                            if val and val not in _existing_kws:
                                _sug_categories["💼 職種・職域"].setdefault(val, []).append(c["name"])
                        # 資格
                        for s in tags.get("certifications", []):
                            if s not in _existing_kws:
                                _sug_categories["📜 資格"].setdefault(s, []).append(c["name"])
                        # 業界
                        for s in tags.get("industries", []):
                            if s not in _existing_kws:
                                _sug_categories["🏢 業界"].setdefault(s, []).append(c["name"])

                    _has_any = any(v for v in _sug_categories.values())
                    if _has_any:
                        # 全選択
                        _all_sug_kws = []
                        for cat_kws in _sug_categories.values():
                            _all_sug_kws.extend(cat_kws.keys())
                        _select_all_cand = st.checkbox(
                            f"すべて選択（{len(_all_sug_kws)}件）", key="dm_sug_select_all"
                        )

                        _selected_sugs = []
                        for cat_name, cat_kws in _sug_categories.items():
                            if not cat_kws:
                                continue
                            st.markdown(f"**{cat_name}**（{len(cat_kws)}件）")
                            _sorted = sorted(cat_kws.items(), key=lambda x: len(x[1]), reverse=True)
                            for sug_kw, cand_names in _sorted:
                                _unique_names = list(set(cand_names))[:3]
                                label = f"{sug_kw} ← {', '.join(_unique_names)}"
                                checked = _select_all_cand or st.session_state.get(f"dm_sug_{cat_name}_{sug_kw}", False)
                                if st.checkbox(label, value=checked, key=f"dm_sug_{cat_name}_{sug_kw}"):
                                    _selected_sugs.append(sug_kw)
                        if _selected_sugs:
                            st.markdown("---")
                            st.markdown(f"**追加されるキーワード（{len(_selected_sugs)}件）:** " +
                                        ", ".join(f"「{k}」" for k in _selected_sugs))
                            _auto_start_sug = st.checkbox("登録と同時に求人取得を開始する", value=True, key="dm_auto_fetch_sug")
                            if st.button("選択したキーワードを登録", type="primary", key="dm_sug_add"):
                                _added_sug_kws = []
                                for skw in _selected_sugs:
                                    if add_keyword(skw, ""):
                                        _added_sug_kws.append(skw)
                                st.success(f"{len(_added_sug_kws)}件のキーワードを追加しました")
                                if _auto_start_sug and _added_sug_kws:
                                    with st.status("求人取得中...", expanded=True) as _sc:
                                        run_fetch_sync(_added_sug_kws, "", list(SOURCE_NAMES), status_container=_sc)
                                    set_app_setting("last_auto_fetch_at", _now_jst().isoformat())
                                st.rerun()
                    else:
                        st.info("追加可能なキーワード候補はありません（全て登録済み）")
                else:
                    st.info("候補者が登録されていません。先に候補者を登録してください。")

            st.markdown("---")
            st.markdown("**登録済みキーワード一覧**")
            st.caption("💡 キーワードを削除すると、そのキーワードに関連する求人データも削除されます。")
            all_kws = get_keywords()

            if all_kws:
                # ステータス集計
                _kw_active = sum(1 for k in all_kws if k.get("enabled"))
                _kw_done = sum(1 for k in all_kws if k.get("fetch_status") == "done")
                _kw_pending = sum(1 for k in all_kws if k.get("fetch_status") in ("pending", "", None))
                st.caption(f"合計 {len(all_kws)}件 ｜ 自動取得ON: {_kw_active} ｜ 取得済み: {_kw_done} ｜ 未取得: {_kw_pending}")

            for kw in all_kws:
                _fs = kw.get("fetch_status", "pending") or "pending"
                _jf = kw.get("jobs_found", "0") or "0"
                _lf = kw.get("last_fetched_at", "") or ""
                _loc = kw.get("location", "")
                _enabled = kw.get("enabled", 1)

                # ステータスバッジ
                if _fs == "fetching":
                    _badge = '<span style="background:#fff3cd;color:#856404;padding:2px 8px;border-radius:10px;font-size:0.75em;">🔄 取得待ち</span>'
                elif _fs == "done":
                    _lf_short = _lf[:16].replace("T", " ") if _lf else ""
                    _badge = f'<span style="background:#d4edda;color:#155724;padding:2px 8px;border-radius:10px;font-size:0.75em;">✅ {_jf}件保存済み</span>'
                elif _fs == "error":
                    _badge = '<span style="background:#f8d7da;color:#721c24;padding:2px 8px;border-radius:10px;font-size:0.75em;">❌ エラー</span>'
                else:
                    _badge = '<span style="background:#e2e3e5;color:#383d41;padding:2px 8px;border-radius:10px;font-size:0.75em;">⏳ 未取得</span>'

                # 有効/無効バッジ
                if not _enabled:
                    _toggle_badge = '<span style="background:#f0f0f0;color:#999;padding:2px 8px;border-radius:10px;font-size:0.75em;">自動取得OFF</span>'
                else:
                    _toggle_badge = ""

                # 勤務地表示
                _loc_str = f"📍{_loc}" if _loc else "📍全国"

                kc1, kc2, kc3, kc4 = st.columns([5, 1.5, 1.5, 0.5])
                kc1.markdown(
                    f"**{esc(kw['keyword'])}** "
                    f"<span style='color:#888;font-size:0.85em;'>{_loc_str}</span><br>"
                    f"{_badge} {_toggle_badge}"
                    + (f" <span style='color:#aaa;font-size:0.7em;'>{_lf[:16].replace('T', ' ')}</span>" if _lf and _fs == "done" else ""),
                    unsafe_allow_html=True
                )
                # 今すぐ取得ボタン
                if kc2.button("▶ 取得", key=f"dm_now_kw_{kw['id']}", help="このキーワードの求人を今すぐ取得"):
                    with st.status(f"「{kw['keyword']}」の求人を取得中...", expanded=True) as _sc:
                        run_fetch_sync([kw["keyword"]], kw.get("location", ""), list(SOURCE_NAMES), status_container=_sc)
                    set_app_setting("last_auto_fetch_at", _now_jst().isoformat())
                    st.rerun()
                # 自動取得ON/OFF
                if _enabled:
                    if kc3.button("自動OFF", key=f"dm_tog_kw_{kw['id']}", help="自動取得の対象外にする"):
                        toggle_keyword(kw["id"], False)
                        st.rerun()
                else:
                    if kc3.button("自動ON", key=f"dm_tog_kw_{kw['id']}", type="primary", help="自動取得の対象に戻す"):
                        toggle_keyword(kw["id"], True)
                        st.rerun()
                if kc4.button("🗑️", key=f"dm_del_kw_{kw['id']}", help="削除"):
                    remove_keyword(kw["id"])
                    st.rerun()

        fetch_loc_sels = st.multiselect("取得勤務地（複数選択可・OR条件）", _LOCATION_OPTIONS, default=["全国"], key="dm_fetch_loc")
        fetch_loc = "" if (not fetch_loc_sels or "全国" in fetch_loc_sels) else fetch_loc_sels[0]

        # 自動更新状態の表示
        if _last_auto:
            try:
                _disp_dt = datetime.fromisoformat(_last_auto)
                from datetime import timedelta as _td
                _next_dt = _disp_dt + _td(hours=_AUTO_REFRESH_HOURS)
                st.caption(f"⏰ 前回の自動取得: {_disp_dt.strftime('%Y-%m-%d %H:%M')} ／ 次回: {_next_dt.strftime('%Y-%m-%d %H:%M')}頃")
            except (ValueError, TypeError):
                pass

        # メイン取得ボタン
        if st.button("🔄 Web求人を自動取得", type="primary", use_container_width=True, key="dm_fetch"):
            kw_list = [kw["keyword"] for kw in registered_kws]
            if not kw_list:
                st.error("キーワードを登録してください")
            elif not enabled_sources:
                st.error("ソースを1つ以上選択してください")
            else:
                with st.status("求人取得中...", expanded=True) as _sc:
                    run_fetch_sync(kw_list, fetch_loc, enabled_sources, status_container=_sc)
                set_app_setting("last_auto_fetch_at", _now_jst().isoformat())
                st.rerun()

        # 半日自動更新（ページ読み込み時に自動実行）
        if _auto_refresh_needed:
            _auto_kw_list = [kw["keyword"] for kw in _auto_kws]
            _auto_sources = list(SOURCE_NAMES)
            with st.status("半日ごとの自動更新中...", expanded=True) as _sc:
                run_fetch_sync(_auto_kw_list, "", _auto_sources, status_container=_sc)
            set_app_setting("last_auto_fetch_at", _now_jst().isoformat())
            st.rerun()

        logs = get_collection_logs(5)
        if logs:
            with st.expander("取得ログ"):
                for log in logs:
                    ran = log.get("ran_at", "")[:16].replace("T", " ")
                    st.caption(f"{ran} | {log.get('sources','')} | 取得:{log.get('jobs_found',0)}/保存:{log.get('jobs_saved',0)} | {log.get('duration_sec',0):.1f}秒")

    # --- 契約中求人登録 ---
    with dm_tabs[1]:
        st.markdown("### 契約中求人の登録")
        st.caption("現在契約中の求人を登録します。「📌 契約中」として区別して管理されます。")

        imp = st.radio("登録方法", ["CSV/Excel", "1件ずつ登録"], horizontal=True, key="dm_c_imp")
        if imp == "CSV/Excel":
            st.caption("カラム: 求人タイトル, 企業名, 勤務地, 年収, URL, 説明, ソース")
            uploaded = st.file_uploader("ファイル", type=["csv", "xlsx", "xls"], key="dm_c_csv")
            if uploaded:
                try:
                    if uploaded.name.endswith(".csv"):
                        content = uploaded.read().decode("utf-8-sig")
                        jobs = parse_csv_upload(content)
                    else:
                        udf = pd.read_excel(uploaded)
                        jobs = parse_csv_upload(udf.to_csv(index=False))
                    if jobs:
                        st.write(f"**{len(jobs)}件**検出")
                        if st.button("契約中求人として登録", type="primary", key="dm_c_imp_btn"):
                            saved = save_jobs(jobs, job_type="contracted")
                            st.success(f"{saved}件を契約中求人として保存")
                            st.rerun()
                except Exception as e:
                    st.error(f"エラー: {e}")
        else:
            with st.form("dm_c_single"):
                j_title = st.text_input("求人タイトル *")
                jc1, jc2 = st.columns(2)
                j_company = jc1.text_input("企業名")
                j_url = jc2.text_input("求人URL")
                jc3, jc4 = st.columns(2)
                j_location = jc3.text_input("勤務地")
                j_salary = jc4.text_input("年収")
                j_desc = st.text_area("説明", height=60)
                if st.form_submit_button("契約中求人として登録", type="primary"):
                    if j_title:
                        save_jobs([{
                            "title": j_title, "company": j_company,
                            "url": j_url or f"contracted_{_now_jst().isoformat()}",
                            "location": j_location, "salary": j_salary,
                            "description": j_desc, "source": "手動登録",
                        }], job_type="contracted")
                        st.success("契約中求人として登録しました")
                        st.rerun()

        st.markdown("---")
        st.markdown("### 登録済み契約中求人")
        contracted_jobs = get_all_jobs(limit=100, job_type="contracted")
        if contracted_jobs:
            for cj in contracted_jobs:
                st.markdown(f"""
                <div class="job-card">
                    📌 <strong>{esc(cj.get('title',''))}</strong>
                    &nbsp;|&nbsp; 🏢 {esc(cj.get('company',''))}
                    &nbsp;|&nbsp; 📍 {esc(cj.get('location',''))}
                    &nbsp;|&nbsp; 💰 {esc(cj.get('salary',''))}
                </div>""", unsafe_allow_html=True)
        else:
            st.info("契約中求人はまだありません")

    # --- 候補者登録 ---
    with dm_tabs[2]:
        st.markdown("### 候補者を登録")
        ext_list = list(SUPPORTED_EXTENSIONS.keys())
        st.caption(f"対応形式: {', '.join(ext_list)}  |  複数ファイル同時アップロード対応（履歴書・職務経歴書・PF等）")

        up_files = st.file_uploader(
            "ファイルをアップロード（複数選択可）",
            type=[e.lstrip(".") for e in ext_list],
            key="cm_upload",
            accept_multiple_files=True,
        )
        if up_files:
            # 各ファイルを読み込み
            parsed_files = []
            file_metas = []
            for uf in up_files:
                file_bytes = uf.read()
                file_text = ""
                cand_data = load_candidate_upload(file_bytes, uf.name)
                if cand_data:
                    doc_type = _detect_file_type(uf.name, " ".join(str(v) for v in cand_data.get("info", {}).values()))
                    cand_data["_file_type"] = doc_type
                    parsed_files.append(cand_data)
                    file_metas.append({
                        "name": uf.name,
                        "type": os.path.splitext(uf.name)[1],
                        "doc_type": doc_type,
                        "size": uf.size if hasattr(uf, "size") else len(file_bytes),
                    })

            if parsed_files:
                # 複数ファイルを統合
                merged = merge_candidate_uploads(parsed_files) if len(parsed_files) > 1 else parsed_files[0]
                info = merged.get("info", {})
                conds = merged.get("conditions", {})
                tags = merged.get("tags", {})

                # ファイル情報表示
                st.success(f"{len(parsed_files)}件のファイルを読み取り完了")
                file_labels = []
                for fm in file_metas:
                    file_labels.append(f"📄 {fm['name']}（{fm['doc_type']}）")
                st.caption(" / ".join(file_labels))

                # 抽出結果を表示
                _cm_c1, _cm_c2 = st.columns(2)
                with _cm_c1:
                    st.markdown("**基本情報**")
                    display_info = {k: v for k, v in info.items()
                                    if k not in ("hard_skills", "soft_skills", "match_reasons",
                                                 "market_score", "market_reasons", "career_summary",
                                                 "personality_memo", "negative_checks", "certifications",
                                                 "industries", "languages", "management", "experience_level",
                                                 "achievements", "education", "work_styles", "availability",
                                                 "career_change_reasons")}
                    for k, v in list(display_info.items())[:10]:
                        st.markdown(f"- {k}: {v}")

                with _cm_c2:
                    st.markdown("**抽出タグ**")
                    # スキル
                    skills = tags.get("skills", [])
                    if skills:
                        st.markdown("🔧 " + ", ".join(f'`{s}`' for s in skills[:8]))
                    # 資格
                    certs = tags.get("certifications", [])
                    if certs:
                        st.markdown("📜 " + ", ".join(f'`{c}`' for c in certs[:5]))
                    # 業界
                    industries = tags.get("industries", [])
                    if industries:
                        st.markdown("🏢 " + ", ".join(f'`{i}`' for i in industries[:5]))
                    # 語学
                    langs = tags.get("languages", [])
                    if langs:
                        st.markdown("🌐 " + ", ".join(f'`{l["language"]}({l["level"]})`' for l in langs[:3]))
                    # 経験レベル
                    exp_level = tags.get("experience_level", "")
                    if exp_level:
                        st.markdown(f"📊 レベル: `{exp_level}`")
                    # マネジメント
                    mgmt = tags.get("management", {})
                    if mgmt.get("has_experience"):
                        team = mgmt.get("team_size", 0)
                        st.markdown(f"👥 マネジメント: `{'{}名規模'.format(team) if team else 'あり'}`")
                    # 面談タグ
                    itags = tags.get("interview_tags", [])
                    if itags:
                        st.markdown("💡 " + ", ".join(f'`{t}`' for t in itags[:6]))
                    # 実績
                    achievements = tags.get("achievements", [])
                    if achievements:
                        st.markdown(f"🏆 実績: {len(achievements)}件検出")
                    # 入社可能時期
                    avail = tags.get("availability", "")
                    if avail:
                        st.markdown(f"📅 入社可能: `{avail}`")
                    # 働き方
                    ws = tags.get("work_styles", [])
                    if ws:
                        st.markdown("🏠 " + ", ".join(f'`{w}`' for w in ws))

                    if not any([skills, certs, industries, langs, itags]):
                        kw_list = conds.get("keywords", [])
                        st.markdown(", ".join(f'`{k}`' for k in kw_list) or "なし")

                save_name = st.text_input("候補者名", value=up_files[0].name.rsplit(".", 1)[0], key="cm_name")
                if st.button("保存", type="primary", key="cm_save"):
                    source_files = merged.get("source_files", [{"name": f.name, "type": _detect_file_type(f.name)} for f in up_files])
                    cid_new = save_candidate(
                        save_name, info, merged.get("strengths", []), conds,
                        tags=tags, source_files=source_files
                    )
                    # ファイル記録
                    for fm in file_metas:
                        save_candidate_file(cid_new, fm["name"], fm["type"], fm["doc_type"], fm.get("size", 0))
                    # 自動プロフィール充実化
                    tmp_cand = {"name": save_name, "info": info, "strengths": merged.get("strengths", []),
                                "conditions": conds, "tags": tags, "id": cid_new}
                    enriched = generate_candidate_profile(tmp_cand)
                    enriched_info = {**info}
                    for key in ["hard_skills", "soft_skills", "match_reasons", "market_score",
                                "market_reasons", "career_summary", "personality_memo",
                                "certifications", "industries", "languages", "management",
                                "experience_level", "achievements", "education", "work_styles",
                                "availability", "career_change_reasons"]:
                        if enriched.get(key):
                            enriched_info[key] = enriched[key]
                    update_candidate(cid_new, info=enriched_info, tags=tags, source_files=source_files)
                    st.success(f"「{save_name}」を保存し、{len(file_metas)}ファイルからプロフィールを自動生成しました")
                    st.rerun()
            else:
                st.warning("ファイルの読み取りに失敗しました。対応形式を確認してください。")

        csv_cands = load_all_candidates()
        if csv_cands:
            st.markdown("---")
            st.markdown("### CSVから一括取り込み")
            if st.button("CSVの候補者をすべて保存", key="cm_bulk"):
                added = 0
                for c in csv_cands:
                    c_tags = c.get("tags", {})
                    save_candidate(c.get("display_name", "候補者"), c.get("info", {}),
                                   c.get("strengths", []), c.get("conditions", {}),
                                   tags=c_tags)
                    added += 1
                st.success(f"{added}名を保存しました")
                st.rerun()

        st.markdown("---")
        st.markdown("### 保存済み候補者")
        if saved_cands:
            for cand in saved_cands:
                mf = evaluate_market_fit(cand)
                star_html = " ⭐" if mf["has_star"] else ""
                st.markdown(f"""
                <div class="cand-card">
                    <strong>{esc(cand.get('name',''))}</strong>{star_html}
                    <span style="color:#a0aec0;margin-left:1rem;font-size:0.8rem;">{esc(cand.get('created_at','')[:10])}</span>
                </div>""", unsafe_allow_html=True)
                cm_c1, cm_c2 = st.columns([1, 1])
                if cm_c1.button("👤 詳細", key=f"cm_pop_{cand['id']}"):
                    show_candidate_popup(cand)
                if cm_c2.button("🗑️ 削除", key=f"cm_del_{cand['id']}"):
                    delete_candidate(cand["id"])
                    st.rerun()
        else:
            st.info("候補者がまだ登録されていません")

    # --- データ管理 ---
    with dm_tabs[3]:
        st.markdown("### データ管理")
        if stats["total_jobs"] > 0:
            if stats.get("sources"):
                st.dataframe(
                    pd.DataFrame([{"ソース": s, "件数": c} for s, c in stats["sources"].items()]),
                    hide_index=True, use_container_width=True,
                )
            mc1, mc2 = st.columns(2)
            if mc1.button("60日以上古いデータを削除", key="dm_old"):
                deleted = delete_old_jobs(60)
                st.info(f"{deleted}件削除")
                st.rerun()
            if mc2.button("全データ削除", key="dm_clear"):
                if st.session_state.get("confirm_clear"):
                    clear_all()
                    st.session_state["confirm_clear"] = False
                    st.rerun()
                else:
                    st.session_state["confirm_clear"] = True
                    st.warning("もう一度クリックで実行")

        with st.expander("📋 取得ログ（デバッグ用）"):
            _fl = get_fetch_log()
            if _fl:
                st.code("\n".join(_fl[-50:]), language="text")
            else:
                st.caption("取得ログはまだありません")

        with st.expander("⚙️ 設定"):
            st.markdown("**スコア重み**")
            defaults_w = {"job": 35, "skill": 35, "soft": 20, "conditions": 10}
            saved_w = get_app_setting("score_weights", defaults_w)
            sw1, sw2 = st.columns(2)
            w_job = sw1.number_input("職種一致度", 0, 100, saved_w.get("job", 35), step=5, key="sw_job")
            w_skill = sw2.number_input("スキル一致度", 0, 100, saved_w.get("skill", 35), step=5, key="sw_skill")
            sw3, sw4 = st.columns(2)
            w_soft = sw3.number_input("ソフトスキル", 0, 100, saved_w.get("soft", 20), step=5, key="sw_soft")
            w_cond = sw4.number_input("条件一致度", 0, 100, saved_w.get("conditions", 10), step=5, key="sw_cond")
            if st.button("重みを保存", key="sw_save"):
                set_app_setting("score_weights", {"job": w_job, "skill": w_skill, "soft": w_soft, "conditions": w_cond})
                st.success("保存しました")

            st.markdown("**ラベル閾値**")
            defaults_l = {"recommended": 85, "good": 70}
            saved_l = get_app_setting("label_thresholds", defaults_l)
            lt1, lt2 = st.columns(2)
            th_rec = lt1.number_input("かなり相性が良い", 50, 100, saved_l.get("recommended", 85), step=5, key="lt_rec")
            th_good = lt2.number_input("相性が良い", 30, 100, saved_l.get("good", 70), step=5, key="lt_good")
            if st.button("閾値を保存", key="lt_save"):
                set_app_setting("label_thresholds", {"recommended": th_rec, "good": th_good})
                st.success("保存しました")
